import os
import base64
import sqlite3
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import customtkinter as ctk
import pandas as pd
import threading
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from core.discord_handler import COLOR_DANGER, DiscordNotifier, COLOR_SUCCESS, COLOR_INFO, COLOR_WARNING
from core.cloud_sync import logica_subir_a_nube, logica_bajar_de_nube
from core.database import (
    actualizar_pedido_db, actualizar_booster_db, actualizar_inventario_db,
    agregar_booster, eliminar_booster, marcar_tarea_completada, obtener_boosters_db,
    realizar_backup_db, obtener_config_precios, actualizar_precio_db,
    agregar_precio_db, eliminar_precio_db, inicializar_db, conectar,
    obtener_conteo_pedidos_activos, obtener_conteo_stock, obtener_ganancia_proyectada,
    finalizar_pedido_db, registrar_abandono_db, obtener_datos_reporte_avanzado,
    obtener_pedidos_activos, crear_pedido, guardar_config_sistema, obtener_config_sistema,
    obtener_kpis_mensuales, registrar_log, obtener_logs_db, obtener_ranking_staff_db,
    obtener_pedidos_mes_actual_db, liquidar_pagos_booster_db, obtener_saldos_pendientes_db,
    obtener_balance_general_db, obtener_historial_completo, obtener_profit_diario_db,
    obtener_total_bote_ranking, obtener_ranking_staff_db, obtener_resumen_mensual_db,
    obtener_resumen_financiero_real, ya_se_ejecuto_hoy
)

from core.logic import (
    calcular_tiempo_transcurrido, calcular_pago_real, calcular_fecha_limite_sugerida
)
from modules.inventario import (
    obtener_inventario_visual, registrar_cuenta_gui, registrar_lote_gui, eliminar_cuenta_gui
)
from modules.pedidos import (
    obtener_elos_en_stock, obtener_cuentas_filtradas_datos
)

# =============================================================================
# CLASE PRINCIPAL: PEREZBOOST APP
# =============================================================================

class PerezBoostApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        inicializar_db()
        self.version = os.getenv("APP_VERSION", "V.Unknown")

        try:
            from core.cloud_sync import MODO_DESARROLLO
        except ImportError:
            MODO_DESARROLLO = True

        if MODO_DESARROLLO:
            env_tag = " [TEST]"
            self.color_status = "#e67e22"
        else:
            env_tag = " [PRODUCCIÓN]"
            self.color_status = "#2ecc71"

        self.title(f"PerezBoost Manager {self.version}{env_tag}")

        self.geometry("1240x750")
        ctk.set_appearance_mode("dark")
        self.centrar_ventana(self, 1240, 750)

        self.menu_contextual = None
        self.tabla_inv = None
        self.tabla_pedidos = None
        self.tabla_boosters = None
        self.tabla_precios = None
        self.tabla_historial = None
        self.tabla_rep = None

        self.datos_inventario = []
        self.map_c_id = {}
        self.map_c_note = {}

        self.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.sidebar_frame = ctk.CTkFrame(self, width=200, corner_radius=0, fg_color="#1a1a1a")
        self.sidebar_frame.grid(row=0, column=0, sticky="nsew")

        self.logo_label = ctk.CTkLabel(self.sidebar_frame, text="🚀 PEREZBOOST", 
                                     font=ctk.CTkFont(size=22, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=30)

        self.crear_boton_menu("🏠 Dashboard", self.mostrar_dashboard, 1)
        self.crear_boton_menu("🤖 Auto-Pilot", self.mostrar_autopilot, 2)
        self.crear_boton_menu("⚙️ Tarifas", self.mostrar_precios, 3)
        self.crear_boton_menu("👥 Boosters", self.mostrar_boosters, 4)
        self.crear_boton_menu("🏆 Ranking Staff", self.mostrar_leaderboard, 5)
        self.crear_boton_menu("📦 Inventario", self.mostrar_inventario, 6)
        self.crear_boton_menu("📜 Pedidos Activos", self.mostrar_pedidos, 7)
        self.crear_boton_menu("💰 Finanzas", self.mostrar_finanzas, 8)
        self.crear_boton_menu("📊 Historial", self.mostrar_historial, 9)
        self.crear_boton_menu("📈 Reportes Pro", self.mostrar_reportes, 10)

        self.content_frame = ctk.CTkFrame(self, corner_radius=15, fg_color="#121212")
        self.content_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")
        self.configurar_menus()
        self.mostrar_dashboard()
        self.iniciar_centinela_automatico()

    # =========================================================================
    # UTILIDADES Y CONFIGURACIÓN UI
    # =========================================================================

    def limpiar_pantalla(self):
        """Elimina widgets y resetea referencias para evitar errores de 'command name'"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        self.tabla_pedidos = None
        self.tabla_historial = None
        self.tabla_inv = None
        self.tabla_boosters = None
        self.tabla_precios = None
        self.tabla_rep = None

    def obtener_bonos_actuales(self):
        try: bp = float(obtener_config_sistema("bono_pedido") or 1.0)
        except: bp = 1.0
        try: bw = float(obtener_config_sistema("bono_wr") or 1.0)
        except: bw = 1.0
        return bp, bw

    def centrar_ventana(self, ventana, ancho, alto):
        ventana.update_idletasks()
        x = (ventana.winfo_screenwidth() // 2) - (ancho // 2)
        y = (ventana.winfo_screenheight() // 2) - (alto // 2)
        ventana.geometry(f"{ancho}x{alto}+{x}+{y}")

    def crear_boton_menu(self, texto, comando, fila):
        boton = ctk.CTkButton(self.sidebar_frame, text=texto, command=comando, corner_radius=10, height=40, fg_color="#2b2b2b", hover_color="#3d3d3d")
        boton.grid(row=fila, column=0, padx=20, pady=10, sticky="ew")

    def configurar_menus(self):
        self.menu_contextual = tk.Menu(self, tearoff=0, bg="#1a1a1a", fg="white", activebackground="#1f538d", font=("Segoe UI", 10))
        self.menu_contextual.add_command(label="📋 Copiar Info (Booster)", command=self.copiar_info_booster)
        self.menu_contextual.add_command(label="✅ Finalizar Pedido", command=self.abrir_ventana_finalizar)
        self.menu_contextual.add_command(label="📝 Editar Información", command=self.abrir_ventana_editar_pedido)
        self.menu_contextual.add_command(label="⏳ Extender Tiempo", command=self.abrir_ventana_extender_tiempo)
        self.menu_contextual.add_separator()
        self.menu_contextual.add_command(label="🚫 Reportar Abandono", command=self.abrir_ventana_reportar_abandono)

    def lanzar_menu_contextual(self, event):
       
        if self.tabla_pedidos:
            item_id = self.tabla_pedidos.identify_row(event.y)
            if item_id:
                self.tabla_pedidos.selection_set(item_id)
                self.menu_contextual.post(event.x_root, event.y_root)

    def configurar_estilo_tabla(self):
        """Configura SOLAMENTE el estilo general, sin tocar instancias específicas."""
        style = ttk.Style()
        style.theme_use("clam")
        
        style.configure("Treeview", background="#1e1e1e", foreground="white", fieldbackground="#1e1e1e", borderwidth=0, font=("Segoe UI", 10), rowheight=35)
        
        style.configure("Treeview.Heading", background="#333333", foreground="white", relief="flat", font=("Segoe UI", 11, "bold"))
        
        style.map("Treeview", background=[('selected', '#1f538d')])

    def copiar_info_booster(self):
        if not self.tabla_pedidos: return
        sel = self.tabla_pedidos.selection()
        if not sel: return

        item_id = sel[0]
        val = self.tabla_pedidos.item(item_id)['values']

        id_pedido = str(val[1]) 
        cuenta = val[4] 
        fecha_raw = str(val[6])

        nota_cuenta = "FRESH"
        try:
            conn = conectar()
            cursor = conn.cursor()
            cursor.execute("SELECT notas FROM pedidos WHERE id = ?", (id_pedido,))
            res = cursor.fetchone()
            if res and res[0]:
                nota_cuenta = res[0]
            conn.close()
        except Exception as e:
            print(f"Error buscando nota: {e}")
        token_raw = f"PB-{id_pedido}".encode('utf-8')
        token_seguro = base64.urlsafe_b64encode(token_raw).decode('utf-8')
        fecha_bonita = fecha_raw
        try:
            meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            fecha_limpia = fecha_raw.split(' ')[0]
            dt = None
            if "-" in fecha_limpia:
                dt = datetime.strptime(fecha_limpia, "%Y-%m-%d")
            elif "/" in fecha_limpia:
                dt = datetime.strptime(fecha_limpia, "%d/%m/%Y")
            if dt:
                fecha_bonita = f"{dt.day} {meses[dt.month - 1]}"
        except Exception as e:
            print(f"Error formateando fecha: {e}")
        URL_DASHBOARD = "https://perezboost-manager.streamlit.app"
        texto_final = f"{cuenta} [{nota_cuenta.upper()}] - Límite: {fecha_bonita} - {URL_DASHBOARD}/?t={token_seguro}"
        self.clipboard_clear()
        self.clipboard_append(texto_final)
        self.update() 
        print(f"✅ Copiado exitoso: {texto_final}")
        messagebox.showinfo("Copiado", f"Info copiada al portapapeles:\n\n{texto_final}")
        
    # =========================================================================
    # SECCIÓN: DASHBOARD
    # =========================================================================

    def mostrar_dashboard(self):
        self.limpiar_pantalla()

        n_pedidos = obtener_conteo_pedidos_activos()
        n_stock = obtener_conteo_stock()
        proyeccion = obtener_ganancia_proyectada()

        profit_neto_hoy = obtener_profit_diario_db()

        ganancia_mes, terminados_mes = obtener_kpis_mensuales()
        criticos, proximos = self.calcular_resumen_emergencias()
        stats_completa = obtener_resumen_mensual_db()
        eficiencia_real = stats_completa[4]

        main_wrapper = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        main_wrapper.pack(expand=True, fill="both")
        container = ctk.CTkFrame(main_wrapper, fg_color="transparent")
        container.place(relx=0.5, rely=0.5, anchor="center") 

        ctk.CTkLabel(container, text=f"PANEL DE CONTROL - {datetime.now().strftime('%B').upper()}", 
                     font=("Arial", 26, "bold")).pack(pady=(0, 25))

        cards_frame = ctk.CTkFrame(container, fg_color="transparent")
        cards_frame.pack(pady=10)

        col_stock = "#e74c3c" if n_stock == 0 else ("#f1c40f" if n_stock < 5 else "#2ecc71")
        
        self.crear_card(cards_frame, "📦 STOCK", f"{n_stock}", col_stock, 0, 0)
        self.crear_card(cards_frame, "⚔️ EN PROGRESO", f"{n_pedidos}", "#3498db", 0, 1)
        self.crear_card(cards_frame, "✅ TERMINADOS", f"{terminados_mes}", "#9b59b6", 0, 2)

        self.crear_card(cards_frame, "💵 PROFIT HOY", f"${profit_neto_hoy:,.2f}", "#27ae60", 1, 0)
        self.crear_card(cards_frame, "💰 PROYECCIÓN", f"${proyeccion:,.2f}", "#00b894", 1, 1) 
        self.crear_card(cards_frame, "⚡ EFICIENCIA", f"{eficiencia_real:.1f} d", "#e67e22", 1, 2)

        btn_frame = ctk.CTkFrame(container, fg_color="transparent")
        btn_frame.pack(pady=40)
        ctk.CTkButton(btn_frame, text="📊 Ver Reporte Detallado", command=self.abrir_reporte_diario,
                      fg_color="#2c3e50", height=45, width=200).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="🔄 Actualizar", command=self.mostrar_dashboard,
                      fg_color="#1f538d", height=45, width=200).pack(side="left", padx=10)

    def abrir_reporte_diario(self):
        hoy_str = datetime.now().strftime("%Y-%m-%d")
        conn = conectar()
        cursor = conn.cursor()
        
        try:
            cursor.execute("""
                SELECT id, booster_nombre, user_pass, elo_inicial, elo_final, wr,
                       fecha_inicio, fecha_fin_real,
                       pago_cliente, pago_booster, ganancia_empresa 
                FROM pedidos 
                WHERE estado = 'Terminado' AND DATE(fecha_fin_real) = DATE(?)
                ORDER BY fecha_fin_real ASC
            """, (hoy_str,))
            
            pedidos_hoy = cursor.fetchall()
            sum_ventas = 0.0; sum_staff = 0.0; sum_profit_db = 0.0
            total_bote_ranking = 0.0    
            
            reporte = f"📊 REPORTE DE CIERRE DIARIO\n📅 Fecha: {datetime.now().strftime('%d/%m/%Y')}\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            
            if not pedidos_hoy:
                reporte += "       (No se registraron cierres hoy)\n"
            
            def clean_f(val):
                try: return float(str(val).replace('$', '').replace(',', '').strip())
                except: return 0.0

            for i, p in enumerate(pedidos_hoy, start=1):
                p_id, staff, cuenta, e_ini, e_fin, wr, f_ini, f_fin, cobro, pago, ganancia = p
                
                v_cobro = clean_f(cobro)
                v_pago = clean_f(pago)
                v_ganancia = clean_f(ganancia)

                try: wr_val = float(wr) 
                except: wr_val = 0.0

                bote_visual = v_cobro - v_pago - v_ganancia
                
                sum_ventas += v_cobro
                sum_staff += v_pago
                sum_profit_db += v_ganancia
                total_bote_ranking += bote_visual

                reporte += f"📦 ORDEN #{i} | {str(staff).upper()}\n"
                reporte += f"   ├─ 🎮 {e_ini} ➔ {e_fin} ({wr_val}% WR)\n"
                reporte += f"   └─ 💰 Margen: +${v_ganancia:.2f} (Bote: ${bote_visual:.2f})\n" 
                reporte += f"────────────────────────────────────\n"

            reporte += (
                f"\n💵 BALANCE FINAL:\n"
                f"   VENTAS BRUTAS:    ${sum_ventas:,.2f}\n"
                f"   PAGO STAFF:      -${sum_staff:,.2f}\n"
                f"   BOTE RANKING:    -${total_bote_ranking:,.2f}\n"
                f"   --------------------------------\n"
                f"   ✅ PROFIT NETO:   ${sum_profit_db:,.2f}\n\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
            )

            v = ctk.CTkToplevel(self)
            v.title("Reporte Ejecutivo")

            self.centrar_ventana(v, 500, 650)

            v.attributes("-topmost", True)  
            v.focus_force()                 
            v.grab_set()                    

            txt_area = ctk.CTkTextbox(v, width=460, height=500, font=("Consolas", 12))
            txt_area.insert("0.0", reporte)
            txt_area.pack(pady=20)
            txt_area.configure(state="disabled")

            ctk.CTkButton(v, text="📋 Copiar", fg_color="#2ecc71", 
                          command=lambda: [self.clipboard_clear(), self.clipboard_append(reporte), 
                                           messagebox.showinfo("Listo", "Copiado")]).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Error: {e}")
        finally: 
            conn.close()
            
    def crear_card(self, master, titulo, valor, color, fila, columna):
        
        card = ctk.CTkFrame(master, corner_radius=15, fg_color="#1a1a1a", 
                            border_width=2, border_color=color, width=260, height=140)
        card.grid(row=fila, column=columna, padx=15, pady=15)
        

        card.grid_propagate(False)

        inner_frame = ctk.CTkFrame(card, fg_color="transparent")
        inner_frame.place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(inner_frame, text=titulo, font=("Arial", 12, "bold"), text_color=color).pack(pady=(0, 5))
        ctk.CTkLabel(inner_frame, text=valor, font=("Arial", 28, "bold")).pack()
        
    def calcular_resumen_emergencias(self):
        
        criticos = 0
        proximos = 0
        hoy = datetime.now().date()
        
        try:
            pedidos = obtener_pedidos_activos()
            for p in pedidos:
                try:
                    fecha_str = str(p[5]).split(' ')[0].strip()
                    fecha_entrega = datetime.strptime(fecha_str, "%Y-%m-%d").date()
                    dias_diferencia = (fecha_entrega - hoy).days
                    
                    if dias_diferencia <= 1:
                        criticos += 1
                    elif dias_diferencia <= 3:
                        proximos += 1
                except:
                    continue
        except Exception as e:
            print(f"Error calculando resumen: {e}")
            
        return criticos, proximos
    
    # =========================================================================
    # SECCIÓN: AUTOPILOT
    # =========================================================================
    
    def mostrar_autopilot(self):
        self.limpiar_pantalla()
        
        header = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        header.pack(pady=20, padx=30, fill="x")
        ctk.CTkLabel(header, text="🤖 AUTO-PILOT OPS (V14.0)", font=("Arial", 24, "bold"), text_color="#2ecc71").pack(side="left")

        container = ctk.CTkScrollableFrame(self.content_frame, fg_color="#1a1a1a", corner_radius=15)
        container.pack(fill="both", expand=True, padx=30, pady=10)

        f1 = ctk.CTkFrame(container, fg_color="#252525", border_width=1, border_color="#e74c3c")
        f1.pack(fill="x", padx=10, pady=10)
        ctk.CTkLabel(f1, text="🚨 Monitoreo de Alerta Roja", font=("Arial", 16, "bold")).pack(pady=10)
        ctk.CTkLabel(f1, text="Envía alertas a Discord si un pedido vence en menos de 12 horas.", text_color="gray").pack()
        ctk.CTkButton(f1, text="Ejecutar Escaneo Ahora", fg_color="#e74c3c", command=self.check_alertas_criticas).pack(pady=15)

        f2 = ctk.CTkFrame(container, fg_color="#252525")
        f2.pack(fill="x", padx=10, pady=10)
        ctk.CTkLabel(f2, text="💸 Bulk Payout (Tesorero)", font=("Arial", 16, "bold")).pack(pady=10)
        ctk.CTkLabel(f2, text="Genera el reporte de pagos pendientes para liquidar a todo el staff.", text_color="gray").pack()
        ctk.CTkButton(f2, text="Generar Nómina .CSV", fg_color="#3498db", command=self.exportar_nomina_csv).pack(pady=15)
        
        f3 = ctk.CTkFrame(container, fg_color="#252525", border_width=1, border_color="#2ecc71")
        f3.pack(fill="x", padx=10, pady=10)
        ctk.CTkLabel(f3, text="📊 Cierre Ejecutivo (Dashboard)", font=("Arial", 16, "bold")).pack(pady=10)
        ctk.CTkLabel(f3, text="Envía el reporte detallado de todas las cuentas cerradas hoy a Discord.", text_color="gray").pack()
        
        ctk.CTkButton(
            f3, 
            text="🚀 Enviar Reporte de Cierre", 
            fg_color="#2ecc71", 
            hover_color="#27ae60",
            height=40,
            command=self.ejecutar_cierre_diario_discord
        ).pack(pady=15)

    def check_alertas_criticas(self):
        url = obtener_config_sistema("discord_webhook_alertas")
        if not url: 
            messagebox.showerror("Error", "Configura el Webhook en 'Tarifas' primero.")
            return

        alertas_enviadas = 0
        pedidos = obtener_pedidos_activos()
        hoy = datetime.now()

        for p in pedidos:
            try:
                fecha_str = str(p[5]).split(' ')[0]
                limite = None

                for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                    try: 
                        limite = datetime.strptime(fecha_str, fmt)
                        break
                    except: continue

                if limite and limite < (hoy + timedelta(hours=24)):
                    notifier = DiscordNotifier(url)
                    notifier.enviar_notificacion(
                        titulo="🚨 RETRASO CRÍTICO",
                        descripcion=f"El pedido de **{p[1]}** (Staff) vence en menos de 12 horas.",
                        color=COLOR_DANGER 
                    )
                    alertas_enviadas += 1

            except Exception as e:
                print(f"Error procesando pedido: {e}")
                continue
        messagebox.showinfo("Auto-Pilot", f"Escaneo finalizado.\nAlertas enviadas a Discord: {alertas_enviadas}")

    def exportar_nomina_csv(self):
        saldos = obtener_saldos_pendientes_db()
        if not saldos:
            messagebox.showinfo("Nómina", "No hay deudas pendientes.")
            return
        
        df = pd.DataFrame(saldos, columns=["Staff", "Total Deuda", "Pedidos", "Detalle"])
        ruta = filedialog.asksaveasfilename(defaultextension=".csv", initialfile=f"Nomina_PerezBoost_{datetime.now().strftime('%Y-%m-%d')}.csv")
        if ruta:
            df.to_csv(ruta, index=False)
            messagebox.showinfo("Éxito", "Nómina exportada correctamente.")
            
    def ejecutar_cierre_diario_discord(self):
        """Genera el reporte detallado del dashboard y lo envía a Discord"""
        hoy_str = datetime.now().strftime("%Y-%m-%d")
        url = obtener_config_sistema("discord_webhook_alertas")
        
        if not url:
            messagebox.showerror("Error", "Configura el Webhook en 'Tarifas'.")
            return

        conn = conectar()
        cursor = conn.cursor()
        
        try:
            notifier = DiscordNotifier(url)
            
            cursor.execute("""
                SELECT id, booster_nombre, user_pass, elo_inicial, elo_final, wr,
                       fecha_inicio, fecha_fin_real,
                       pago_cliente, pago_booster, ganancia_empresa 
                FROM pedidos 
                WHERE estado = 'Terminado' AND DATE(fecha_fin_real) = DATE(?)
                ORDER BY fecha_fin_real ASC
            """, (hoy_str,))
            
            pedidos_hoy = cursor.fetchall()
            sum_ventas = 0.0; sum_staff = 0.0; sum_profit_db = 0.0
            total_bote_ranking = 0.0    

            reporte = f"📊 **REPORTE EJECUTIVO DE CIERRE**\n📅 **Fecha:** {datetime.now().strftime('%d/%m/%Y')}\n"
            reporte += "━━━━━━━━━━━━━━━━━━━━\n\n"
            
            if not pedidos_hoy:
                reporte += "*(No se registraron cierres hoy)*\n"
            
            def clean_f(val):
                try: return float(str(val).replace('$', '').replace(',', '').strip())
                except: return 0.0

            for i, p in enumerate(pedidos_hoy, start=1):
                p_id, staff, cuenta, e_ini, e_fin, wr, f_ini, f_fin, cobro, pago, ganancia = p
                
                v_cobro = clean_f(cobro)
                v_pago = clean_f(pago)
                v_ganancia = clean_f(ganancia)
                try: wr_val = float(wr) 
                except: wr_val = 0.0

                b_ped, b_wr = self.obtener_bonos_actuales()
                bote_visual = (b_ped + b_wr) if wr_val >= 60 else b_ped

                sum_ventas += v_cobro
                sum_staff += v_pago
                sum_profit_db += v_ganancia
                total_bote_ranking += bote_visual

                reporte += f"📦 **ORDEN #{i}** | {str(staff).upper()}\n"
                reporte += f"├─ 🎮 {e_ini} ➔ {e_fin} ({wr_val}% WR)\n"
                reporte += f"└─ 💰 Margen: `+${v_ganancia:.2f}` (Bote: `${bote_visual:.2f}`)\n" 
                reporte += f"──────────────────\n"

            reporte += (
                f"\n💵 **BALANCE FINAL:**\n"
                f"• Ventas Brutas: `${sum_ventas:,.2f}`\n"
                f"• Pago Staff: `-${sum_staff:,.2f}`\n"
                f"• Bote Ranking: `-${total_bote_ranking:,.2f}`\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"✅ **PROFIT NETO: `${sum_profit_db:,.2f}`**"
            )
            notifier.enviar_notificacion(
                titulo="📅 CIERRE DE CAJA DIARIO",
                descripcion=reporte,
                color=COLOR_SUCCESS
            )
            messagebox.showinfo("Auto-Pilot", "Reporte detallado enviado.")
        except Exception as e:
            messagebox.showerror("Error", f"Fallo al enviar: {e}")

    def iniciar_centinela_automatico(self):
        """Hilo en segundo plano que gestiona el Auto-Pilot"""
        def loop_centinela():
            import time
            while True:
                if not ya_se_ejecuto_hoy("alertas_24h"):
                    print("🤖 Auto-Pilot: Ejecutando chequeo diario de alertas...")
                    self.check_alertas_criticas_silencioso()
                    marcar_tarea_completada("alertas_24h")
                time.sleep(3600)

        threading.Thread(target=loop_centinela, daemon=True).start()

    def check_alertas_criticas_silencioso(self):
        """Versión de alertas que no interrumpe con ventanas emergentes"""
        url = obtener_config_sistema("discord_webhook_alertas") or obtener_config_sistema("discord_webhook")
        if not url: return
        pedidos = obtener_pedidos_activos()
        hoy = datetime.now()
        for p in pedidos:
            try:
                fecha_str = str(p[5]).split(' ')[0]
                limite = None
                for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                    try: limite = datetime.strptime(fecha_str, fmt); break
                    except: continue
                
                if limite and limite < (hoy + timedelta(hours=24)):
                    notifier = DiscordNotifier(url)
                    notifier.enviar_notificacion(
                        titulo="🚨 REVISIÓN DIARIA DE SEGURIDAD",
                        descripcion=f"**Booster:** {p[1]}\n**Cuenta:** {p[4]}\nQuedan menos de 24h para la entrega.",
                        color=COLOR_WARNING
                    )
            except: continue

    # =========================================================================
    # SECCIÓN: TARIFAS Y CONFIGURACIÓN NUBE
    # =========================================================================

    def mostrar_precios(self):
        self.limpiar_pantalla()
        self.configurar_estilo_tabla()

        scroll_container = ctk.CTkScrollableFrame(self.content_frame, fg_color="transparent")
        scroll_container.pack(fill="both", expand=True)

        header = ctk.CTkFrame(scroll_container, fg_color="transparent")
        header.pack(pady=15, padx=30, fill="x")
        ctk.CTkLabel(header, text="⚙️ CONFIGURACIÓN DE SISTEMA", font=("Arial", 20, "bold")).pack(side="left")

        cols = ("div", "p_cli", "m_per", "p_boo", "pts")
        self.tabla_precios = ttk.Treeview(scroll_container, columns=cols, show="headings", height=6)

        headers = ["DIVISIÓN", "PRECIO CLIENTE", "MARGEN PEREZ", "PAGO BOOSTER", "PUNTOS RANK"]
        for col, h in zip(cols, headers):
            self.tabla_precios.heading(col, text=h)
            ancho = 100 if col == "pts" else 150
            self.tabla_precios.column(col, anchor="center", width=ancho)

        self.tabla_precios.pack(padx=30, pady=10, fill="x") 
        self.actualizar_tabla_precios()

        footer = ctk.CTkFrame(scroll_container, fg_color="transparent")
        footer.pack(pady=(5, 10), padx=30, fill="x")
        
        ctk.CTkButton(footer, text="+ Nueva Tarifa", fg_color="#2ecc71", width=110,
                      command=self.abrir_ventana_nuevo_precio).pack(side="left", padx=5)
        ctk.CTkButton(footer, text="📝 Editar", fg_color="#f39c12", width=100,
                      command=self.abrir_ventana_editar_precio).pack(side="left", padx=5)
        
        ctk.CTkFrame(footer, width=20, height=1, fg_color="transparent").pack(side="left") 
        
        ctk.CTkButton(footer, text="💾 Backup Local", fg_color="#1f538d", width=110,
                      command=self.ejecutar_backup_manual).pack(side="left", padx=5)
        ctk.CTkButton(footer, text="🕵️ Auditoría", fg_color="#34495e", hover_color="#2c3e50", width=110,
                      command=self.abrir_visor_logs).pack(side="left", padx=5)
        ctk.CTkButton(footer, text="🗑️ Eliminar", fg_color="#e74c3c", width=100,
                      command=self.eliminar_precio_seleccionado).pack(side="right")

        ctk.CTkFrame(scroll_container, height=2, fg_color="#333333").pack(fill="x", padx=30, pady=10)

        # ==========================================
        # 💰 NUEVO FRAME: CONFIGURACIÓN DEL BOTÍN
        # ==========================================
        frame_botin = ctk.CTkFrame(scroll_container, fg_color="#1a1a1a", corner_radius=10)
        frame_botin.pack(fill="x", padx=30, pady=5)
        
        ctk.CTkLabel(frame_botin, text="💰 CONFIGURACIÓN DE BOTÍN (RANKING)", font=("Arial", 14, "bold"), text_color="#f1c40f").pack(anchor="w", padx=15, pady=(10,5))
        
        row_b1 = ctk.CTkFrame(frame_botin, fg_color="transparent")
        row_b1.pack(fill="x", padx=10, pady=2)
        ctk.CTkLabel(row_b1, text="Bono por Pedido ($):", width=180, anchor="w").pack(side="left")
        self.entry_bono_pedido = ctk.CTkEntry(row_b1, width=100)
        self.entry_bono_pedido.pack(side="left", padx=5)
        
        row_b2 = ctk.CTkFrame(frame_botin, fg_color="transparent")
        row_b2.pack(fill="x", padx=10, pady=2)
        ctk.CTkLabel(row_b2, text="Bono Extra por Alto WR ($):", width=180, anchor="w").pack(side="left")
        self.entry_bono_wr = ctk.CTkEntry(row_b2, width=100)
        self.entry_bono_wr.pack(side="left", padx=5)

        val_ped = obtener_config_sistema("bono_pedido")
        val_wr = obtener_config_sistema("bono_wr")
        self.entry_bono_pedido.insert(0, val_ped if val_ped else "1.0")
        self.entry_bono_wr.insert(0, val_wr if val_wr else "1.0")

        ctk.CTkButton(frame_botin, text="💾 Guardar Botín", width=150, height=30, fg_color="#d35400", hover_color="#e67e22", 
                      command=self.guardar_config_botin).pack(pady=10)

        # ==========================================
        # 🤖 FRAME DISCORD
        # ==========================================
        frame_discord = ctk.CTkFrame(scroll_container, fg_color="#1a1a1a", corner_radius=10)
        frame_discord.pack(fill="x", padx=30, pady=5)
        
        ctk.CTkLabel(frame_discord, text="🤖 CONECTIVIDAD DISCORD", font=("Arial", 14, "bold"), text_color="#5865F2").pack(anchor="w", padx=15, pady=(10,5))

        row1 = ctk.CTkFrame(frame_discord, fg_color="transparent")
        row1.pack(fill="x", padx=10, pady=2)
        ctk.CTkLabel(row1, text="🔔 Canal Pedidos/Log:", width=150, anchor="w").pack(side="left")
        self.entry_webhook = ctk.CTkEntry(row1, width=400, placeholder_text="Webhook General...")
        self.entry_webhook.pack(side="left", fill="x", expand=True, padx=5)

        row2 = ctk.CTkFrame(frame_discord, fg_color="transparent")
        row2.pack(fill="x", padx=10, pady=2)
        ctk.CTkLabel(row2, text="🏆 Canal Ranking:", width=150, anchor="w").pack(side="left")
        self.entry_webhook_rank = ctk.CTkEntry(row2, width=400, placeholder_text="Webhook Ranking...")
        self.entry_webhook_rank.pack(side="left", fill="x", expand=True, padx=5)
        
        row3 = ctk.CTkFrame(frame_discord, fg_color="transparent")
        row3.pack(fill="x", padx=10, pady=2)
        ctk.CTkLabel(row3, text="🚨 Canal Alerta Roja:", width=150, anchor="w").pack(side="left")
        self.entry_webhook_alerts = ctk.CTkEntry(row3, width=400, placeholder_text="Webhook de Alertas Críticas...")
        self.entry_webhook_alerts.pack(side="left", fill="x", expand=True, padx=5)

        url_pedidos = obtener_config_sistema("discord_webhook")
        url_ranking = obtener_config_sistema("discord_webhook_ranking")
        url_alertas = obtener_config_sistema("discord_webhook_alertas")
        if url_alertas: self.entry_webhook_alerts.insert(0, url_alertas)
        if url_pedidos: self.entry_webhook.insert(0, url_pedidos)
        if url_ranking: self.entry_webhook_rank.insert(0, url_ranking)
        
        ctk.CTkButton(frame_discord, text="Guardar Webhooks", width=150, height=30, fg_color="#404eed", 
                      command=self.guardar_webhooks_discord).pack(pady=10)

        # ==========================================
        # ☁️ FRAME CLOUD
        # ==========================================
        frame_cloud = ctk.CTkFrame(scroll_container, fg_color="#2d2042", corner_radius=10) 
        frame_cloud.pack(fill="x", padx=30, pady=10)

        ctk.CTkLabel(frame_cloud, text="☁️ SINCRONIZACIÓN NUBE", font=("Arial", 14, "bold"), text_color="#a29bfe").pack(anchor="w", padx=15, pady=(10,5))

        btns_cloud = ctk.CTkFrame(frame_cloud, fg_color="transparent")
        btns_cloud.pack(pady=10)

        btn_subir = ctk.CTkButton(
            btns_cloud, 
            text="⬆️ SUBIR a Nube\n(Backup)", 
            fg_color="#8e44ad", hover_color="#9b59b6", 
            width=160, height=50,
            font=("Arial", 12, "bold"),
            command=self.accion_subir_nube 
        )
        btn_subir.pack(side="left", padx=20)

        btn_bajar = ctk.CTkButton(
            btns_cloud, 
            text="⬇️ BAJAR de Nube\n(Restore)", 
            fg_color="#e67e22", hover_color="#d35400", 
            width=160, height=50,
            font=("Arial", 12, "bold"),
            command=self.accion_bajar_nube 
        )
        btn_bajar.pack(side="left", padx=20)

        ctk.CTkLabel(frame_cloud, text="Nota: 'Subir' guarda tu PC en la nube. 'Bajar' trae la nube a tu PC (sobrescribe local).", font=("Arial", 10, "italic"), text_color="gray").pack(pady=(0,10))

    def guardar_config_botin(self):
        try:
            val_p = float(self.entry_bono_pedido.get().strip())
            val_w = float(self.entry_bono_wr.get().strip())
            
            guardar_config_sistema("bono_pedido", str(val_p))
            guardar_config_sistema("bono_wr", str(val_w))
            
            messagebox.showinfo("Éxito", "Configuración del botín actualizada correctamente.")
        except ValueError:
            messagebox.showerror("Error", "Debes ingresar valores numéricos válidos (Ej: 1.0, 0.5, 0.0)")

    def guardar_webhooks_discord(self):
        url_pedidos = self.entry_webhook.get().strip()
        url_ranking = self.entry_webhook_rank.get().strip()
        url_alertas = self.entry_webhook_alerts.get().strip()
        
        g1 = guardar_config_sistema("discord_webhook", url_pedidos)
        g2 = guardar_config_sistema("discord_webhook_ranking", url_ranking)
        g3 = guardar_config_sistema("discord_webhook_alertas", url_alertas)
        
        if g1 and g2 and g3:
            messagebox.showinfo("Discord", "✅ Conexiones (Pedidos, Ranking y Alertas) guardadas.")
        else:
            messagebox.showerror("Error", "Hubo un problema guardando la configuración.")

    def actualizar_tabla_precios(self):
        for item in self.tabla_precios.get_children():
            self.tabla_precios.delete(item)

        tarifas = obtener_config_precios()
        
        for t in tarifas:
            div, p_cli, m_per, pts = t
            p_boo = p_cli - m_per
            self.tabla_precios.insert("", "end", values=(
                div,
                f"${p_cli:,.2f}",
                f"${m_per:,.2f}",
                f"${p_boo:,.2f}",
                pts
            ))
            
    def accion_subir_nube(self):
        self.win = ctk.CTkToplevel(self)
        self.centrar_ventana(self.win, 300, 150)
        self.win.attributes("-topmost", True)
        ctk.CTkLabel(self.win, text="🚀 Subiendo...", font=("Arial", 14)).pack(pady=20)
        bar = ctk.CTkProgressBar(self.win, mode="indeterminate")
        bar.pack(pady=10)
        bar.start()
        
        def proceso_subida():
            try:
                import psycopg2
                from dotenv import load_dotenv
                import os
                load_dotenv(".env")
                
                url_db = os.getenv("DATABASE_URL")
                if url_db:

                    conn_pg = psycopg2.connect(url_db)
                    cur_pg = conn_pg.cursor()
                    cur_pg.execute("SELECT id, opgg FROM pedidos WHERE opgg IS NOT NULL AND opgg != ''")
                    links_nube = cur_pg.fetchall()
                    conn_pg.close()

                    if links_nube:
                        conn_local = conectar()
                        cur_local = conn_local.cursor()
                        for pid, link in links_nube:
                            cur_local.execute("UPDATE pedidos SET opgg = ? WHERE id = ?", (link, pid))
                        conn_local.commit()
                        conn_local.close()
                        print(f"🛡️ {len(links_nube)} enlaces OP.GG rescatados con éxito.")
            except Exception as e:
                print(f"⚠️ Aviso: No se pudieron rescatar los OPGG de la nube: {e}")
            def fin(): 
                self.win.destroy()
                messagebox.showinfo("Nube", "✅ Datos subidos correctamente.")
                self.mostrar_tracking() if hasattr(self, 'mostrar_tracking') else None
                
            def err(e): 
                self.win.destroy()
                messagebox.showerror("Error", f"Fallo al subir: {e}")
                
            logica_subir_a_nube(fin, err)
            
        threading.Thread(target=proceso_subida).start()
        
    def accion_bajar_nube(self):
        if not messagebox.askyesno("Confirmar", "⚠️ ¿Borrar datos locales y traer los de la Nube?"): return
        wb_pedidos = obtener_config_sistema("discord_webhook")
        wb_ranking = obtener_config_sistema("discord_webhook_ranking")
        wb_alertas = obtener_config_sistema("discord_webhook_alertas")
        b_ped = obtener_config_sistema("bono_pedido")
        b_wr = obtener_config_sistema("bono_wr")

        self.win = ctk.CTkToplevel(self)
        self.centrar_ventana(self.win, 300, 150)
        self.win.attributes("-topmost", True)
        ctk.CTkLabel(self.win, text="⬇️ Sincronizando con la Nube...", font=("Arial", 14)).pack(pady=20)
        bar = ctk.CTkProgressBar(self.win, mode="indeterminate"); bar.pack(pady=10); bar.start()

        def fin(): 
            if wb_pedidos: guardar_config_sistema("discord_webhook", wb_pedidos)
            if wb_ranking: guardar_config_sistema("discord_webhook_ranking", wb_ranking)
            if wb_alertas: guardar_config_sistema("discord_webhook_alertas", wb_alertas)
            if b_ped: guardar_config_sistema("bono_pedido", b_ped)
            if b_wr: guardar_config_sistema("bono_wr", b_wr)

            self.win.destroy()
            messagebox.showinfo("Nube", "✅ Datos sincronizados correctamente.")
            self.mostrar_dashboard()
            
        def err(e): 
            self.win.destroy()
            messagebox.showerror("Error", f"Fallo: {e}")
            
        threading.Thread(target=logica_bajar_de_nube, args=(fin, err)).start()
        
    # =========================================================================
    # SECCIÓN: BOOSTERS (STAFF)
    # =========================================================================

    def mostrar_boosters(self):
        self.limpiar_pantalla()
        self.configurar_estilo_tabla()
        
        header = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        header.pack(pady=(15, 5), padx=30, fill="x")
        ctk.CTkLabel(header, text="👥 GESTIÓN DE STAFF", font=("Arial", 20, "bold")).pack(side="left")
        
        self.entry_busqueda_b = ctk.CTkEntry(header, placeholder_text="Nombre del Booster...", width=200)
        self.entry_busqueda_b.pack(side="right")
        ctk.CTkButton(header, text="🔍", width=40, command=self.filtrar_boosters).pack(side="right", padx=5)

        cols = ("id_v", "id_r", "nombre")
        self.tabla_boosters = ttk.Treeview(self.content_frame, columns=cols, show="headings")
        self.tabla_boosters.heading("id_v", text="#")
        self.tabla_boosters.heading("nombre", text="NOMBRE DEL STAFF")
        self.tabla_boosters.column("id_v", width=60, anchor="center", stretch=False)
        self.tabla_boosters.column("id_r", width=0, stretch=tk.NO) 
        self.tabla_boosters.column("nombre", width=600, anchor="center") 

        self.tabla_boosters.pack(padx=30, pady=10, fill="both", expand=True)
        self.filtrar_boosters()

        footer = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        footer.pack(pady=(5, 20), padx=30, fill="x")
        ctk.CTkButton(footer, text="+ Nuevo Booster", fg_color="#2ecc71", command=self.abrir_ventana_booster).pack(side="left")
        ctk.CTkButton(footer, text="📝 Editar", fg_color="#f39c12", command=self.abrir_ventana_editar_booster).pack(side="left", padx=5)
        ctk.CTkButton(footer, text="🗑️ Despedir", fg_color="#e74c3c", command=self.eliminar_booster_seleccionado).pack(side="right")

    def filtrar_boosters(self):
        query = self.entry_busqueda_b.get().lower()
        for i in self.tabla_boosters.get_children(): self.tabla_boosters.delete(i)
        for i, b in enumerate(obtener_boosters_db(), start=1):
            if query == "" or query in b[1].lower():
                self.tabla_boosters.insert("", tk.END, values=(i, b[0], b[1]))

    # =========================================================================
    # SECCIÓN: INVENTARIO (STOCK)
    # =========================================================================

    def mostrar_inventario(self):
        self.limpiar_pantalla()
        self.configurar_estilo_tabla()
        
        header = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        header.pack(pady=15, padx=30, fill="x")
        ctk.CTkLabel(header, text="📦 STOCK DISPONIBLE", font=("Arial", 20, "bold")).pack(side="left")
        
        self.entry_busqueda_i = ctk.CTkEntry(header, placeholder_text="Buscar Elo...", width=200)
        self.entry_busqueda_i.pack(side="right")

        self.entry_busqueda_i.bind("<KeyRelease>", lambda event: self.filtrar_inventario())
        ctk.CTkButton(header, text="🔍", width=40, command=self.filtrar_inventario).pack(side="right", padx=5)

        cols = ("id_v", "elo", "user_pass", "desc")
        self.tabla_inv = ttk.Treeview(self.content_frame, columns=cols, show="headings")
        
        headers = ["#", "ELO", "USUARIO / CONTRASEÑA", "NOTAS"]
        anchos = [50, 120, 300, 200]

        for col, txt, ancho in zip(cols, headers, anchos):
            self.tabla_inv.heading(col, text=txt)
            self.tabla_inv.column(col, width=ancho, anchor="center")

        self.tabla_inv.pack(padx=30, pady=10, fill="both", expand=True)
        self.filtrar_inventario()

        footer = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        footer.pack(pady=(5, 20), padx=30, fill="x")
        ctk.CTkButton(footer, text="+ Añadir", fg_color="#2ecc71", command=self.abrir_ventana_registro).pack(side="left", padx=5)
        ctk.CTkButton(footer, text="📝 Editar", fg_color="#f39c12", command=self.abrir_ventana_editar_inventario).pack(side="left", padx=5)
        ctk.CTkButton(footer, text="📥 Masivo", fg_color="#3498db", command=self.abrir_ventana_masivo).pack(side="left")
        ctk.CTkButton(footer, text="🗑️ Borrar", fg_color="#e74c3c", command=self.eliminar_seleccionado).pack(side="right")

    def filtrar_inventario(self):
        query = self.entry_busqueda_i.get().lower()
        for i in self.tabla_inv.get_children(): self.tabla_inv.delete(i)
        
        self.datos_inventario = obtener_inventario_visual()
        
        for d in self.datos_inventario:
            if query == "" or query in str(d[3]).lower() or query in str(d[2]).lower():
                self.tabla_inv.insert("", tk.END, values=(d[0], d[3], d[2], d[4]))

    # =========================================================================
    # SECCIÓN: PEDIDOS ACTIVOS
    # =========================================================================

    def mostrar_pedidos(self):
        self.limpiar_pantalla()
        self.configurar_estilo_tabla()
        
        header = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        header.pack(pady=(15, 5), padx=30, fill="x")
        ctk.CTkLabel(header, text="⚔️ PEDIDOS ACTIVOS", font=("Arial", 20, "bold")).pack(side="left")
        
        self.entry_busqueda = ctk.CTkEntry(header, placeholder_text="Buscar Booster...", width=200)
        self.entry_busqueda.bind("<KeyRelease>", self.filtrar_pedidos)
        self.entry_busqueda.pack(side="right")
        ctk.CTkButton(header, text="🔍", width=40, command=self.filtrar_pedidos).pack(side="right", padx=5)

        cols = ("id_v", "id_r", "booster", "elo", "cuenta", "inicio", "fin", "tiempo") 
        self.tabla_pedidos = ttk.Treeview(self.content_frame, columns=cols, show="headings")
        self.tabla_pedidos.bind("<Button-3>", self.lanzar_menu_contextual)
        
        self.tabla_pedidos.tag_configure('urgente', foreground='#8b0000')
        self.tabla_pedidos.tag_configure('alerta', foreground='#ffa500')
        self.tabla_pedidos.tag_configure('normal', foreground='#2ecc71')
        
        headers = ["#", "ID_R", "STAFF", "ELO", "CUENTA / USER", "INICIO", "FIN", "TIEMPO"]
        anchos = [40, 0, 130, 100, 200, 100, 100, 130]

        for c, h, w in zip(cols, headers, anchos):
            self.tabla_pedidos.heading(c, text=h)
            if c == "id_r":
                self.tabla_pedidos.column(c, width=0, stretch=tk.NO)
            else:
                self.tabla_pedidos.column(c, width=w, anchor="center", stretch=(c == "cuenta"))

        self.tabla_pedidos.pack(padx=30, pady=10, fill="both", expand=True)
        self.filtrar_pedidos()

        footer = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        footer.pack(pady=(5, 20), padx=30, fill="x")
        ctk.CTkButton(footer, text="⚡ Nuevo", fg_color="#3498db", command=self.abrir_ventana_nuevo_pedido).pack(side="left")
        ctk.CTkButton(footer, text="⏳ Extender", command=self.abrir_ventana_extender_tiempo).pack(side="left", padx=5)
        ctk.CTkButton(footer, text="📝 Editar", fg_color="#f39c12", command=self.abrir_ventana_editar_pedido).pack(side="left", padx=5)
        ctk.CTkButton(footer, text="✅ Finalizar", fg_color="#2ecc71", command=self.abrir_ventana_finalizar).pack(side="right")
        ctk.CTkButton(footer, text="🚫 Abandono", fg_color="#e74c3c", command=self.abrir_ventana_reportar_abandono).pack(side="right", padx=10)
        ctk.CTkButton(footer, text="🔨 Ban", fg_color="#8b0000", hover_color="#5a0000", command=self.reportar_ban_seleccionado).pack(side="right")

    def filtrar_pedidos(self, event=None):
        query = self.entry_busqueda.get().lower().strip()
        hoy = datetime.now().date()
        
        if not self.tabla_pedidos: return
        for i in self.tabla_pedidos.get_children(): self.tabla_pedidos.delete(i)
        
        try:
            datos_raw = obtener_pedidos_activos()
            if not datos_raw: return

            for i, p in enumerate(datos_raw, start=1):
                try:
                    def procesar_fecha(txt):
                        txt = str(txt).split(' ')[0].strip()
                        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                            try:
                                return datetime.strptime(txt, fmt).date()
                            except:
                                continue
                        return None

                    dt_inicio = procesar_fecha(p[4])
                    dt_limite = procesar_fecha(p[5])

                    if dt_inicio and dt_limite:
                        dias_activa = (hoy - dt_inicio).days
                        dias_restantes = (dt_limite - hoy).days

                        if dias_restantes < 0: 
                            tag, ico = 'urgente', f"💀 {dias_activa}d"
                        elif dias_restantes <= 1: 
                            tag, ico = 'urgente', f"🔴 {dias_activa}d"
                        elif dias_restantes <= 3: 
                            tag, ico = 'alerta', f"🟡 {dias_activa}d"
                        else: 
                            tag, ico = 'normal', f"🟢 {dias_activa}d"

                        f_ini_v = dt_inicio.strftime("%d/%m/%Y")
                        f_lim_v = dt_limite.strftime("%d/%m/%Y")
                    else:
                        raise ValueError("Formato no reconocido")

                except Exception as e:
                    f_ini_v, f_lim_v, ico, tag = str(p[4]), str(p[5]), "⚪ Error", "normal"

                if query == "" or query in str(p[0]) or query in str(p[1]).lower() or query in str(p[3]).lower():
                    fila = (i, p[0], p[1], p[2], p[3], f_ini_v, f_lim_v, ico)
                    self.tabla_pedidos.insert("", tk.END, values=fila, tags=(tag,))
                    
        except Exception as e:
            print(f"Error en tabla: {e}")

    # =========================================================================
    # SECCIÓN: HISTORIAL
    # =========================================================================

    def mostrar_historial(self):
        self.limpiar_pantalla()
        self.configurar_estilo_tabla()

        header = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        header.pack(pady=15, padx=20, fill="x")
        ctk.CTkLabel(header, text="📊 HISTORIAL OPERATIVO", font=("Arial", 20, "bold")).pack(side="left")

        ctk.CTkButton(header, text="⚠️ Ver Abandonos", fg_color="#e74c3c", hover_color="#c0392b",
                      width=120, command=lambda: self.filtrar_historial(solo_abandonos=True)).pack(side="right", padx=10)

        self.entry_busqueda_h = ctk.CTkEntry(header, placeholder_text="Filtrar Staff/Estado...", width=200)
        self.entry_busqueda_h.pack(side="right")
        self.entry_busqueda_h.bind("<KeyRelease>", lambda e: self.filtrar_historial())
        ctk.CTkButton(header, text="🔍", width=40, command=self.filtrar_historial).pack(side="right", padx=5)

        cols = ("id_visual", "booster", "cuenta", "wr", "inicio", "fin", "duracion", "estado_oculto")
        self.tabla_historial = ttk.Treeview(self.content_frame, columns=cols, show="headings")

        anchos = [50, 150, 280, 70, 110, 110, 100, 0] 
        headers = ["#", "STAFF", "CUENTA / ELO FINAL", "WR", "INICIO", "FIN", "DURACIÓN", ""]
        
        for col, head, ancho in zip(cols, headers, anchos):
            self.tabla_historial.heading(col, text=head, command=lambda c=col: self.ordenar_columna(c, False))
            
            if col == "estado_oculto": 
                self.tabla_historial.column(col, width=0, stretch=tk.NO)
            elif col == "id_visual":
                self.tabla_historial.column(col, width=ancho, anchor="center", stretch=tk.NO)
            else: 
                self.tabla_historial.column(col, width=ancho, anchor="center", stretch=(col=="cuenta"))

        self.tabla_historial.tag_configure('terminado', foreground='#2ecc71')
        self.tabla_historial.tag_configure('abandonado', foreground='#e74c3c')
        self.tabla_historial.tag_configure('baneada', foreground='#8b0000')
        self.tabla_historial.pack(padx=20, pady=10, fill="both", expand=True)

        self.panel_total_h = ctk.CTkFrame(self.content_frame, fg_color="#1a1a1a", corner_radius=10, height=40)
        self.panel_total_h.pack(pady=10, fill="x", padx=20)
        self.panel_total_h.pack_propagate(False)
        self.lbl_totales_h = ctk.CTkLabel(self.panel_total_h, text="", font=("Arial", 14, "bold"), text_color="#7f8c8d")
        self.lbl_totales_h.pack(expand=True)
        
        self.filtrar_historial()
        
    def ordenar_columna(self, col, reverse):
        """Ordena las filas del Treeview al hacer clic en el encabezado"""
        l = [(self.tabla_historial.set(k, col), k) for k in self.tabla_historial.get_children('')]

        try:

            if col in ("inicio", "fin"):
                def convertir_fecha(txt):
                    try:

                        txt = str(txt).split(' ')[0].strip()
                        return datetime.strptime(txt, "%d/%m/%Y")
                    except:
                        return datetime.min
                
                l.sort(key=lambda t: convertir_fecha(t[0]), reverse=reverse)

            elif col == "id_visual":
                 l.sort(key=lambda t: int(t[0]), reverse=reverse)

            else:
                l.sort(reverse=reverse)
        except Exception as e:
            print(f"Error ordenando: {e}")
            l.sort(reverse=reverse)

        for index, (val, k) in enumerate(l):
            self.tabla_historial.move(k, '', index)
        self.tabla_historial.heading(col, command=lambda: self.ordenar_columna(col, not reverse))
        
    def filtrar_historial(self, solo_abandonos=False):
        query = self.entry_busqueda_h.get().lower()
        if not self.tabla_historial: return

        for item in self.tabla_historial.get_children():
            self.tabla_historial.delete(item)
        
        try:
            datos = obtener_historial_completo()
            contador_visual = 1
            
            for row in datos:

                id_ped_real, booster, user, elo_fin, ini_raw, fin_raw, estado, wr = row
                est_str = str(estado).upper()

                mostrar = False
                if solo_abandonos:
                    if "ABANDONADO" in est_str: mostrar = True
                else:
                    if query == "" or query in str(booster).lower() or query in est_str.lower(): 
                        mostrar = True
                
                if mostrar:

                    def f_v(t):
                        if not t: return "---"
                        t = str(t).split(' ')[0]
                        try: return datetime.strptime(t, "%Y-%m-%d").strftime("%d/%m/%Y")
                        except: return t

                    duracion = "---"

                    try:
                        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                            try:
                                f1 = datetime.strptime(str(ini_raw).split(' ')[0], fmt)
                                f2 = datetime.strptime(str(fin_raw).split(' ')[0], fmt)
                                d = (f2 - f1).days
                                duracion = f"{d} días" if d > 0 else "⚡ Hoy"
                                break
                            except: continue
                    except: pass

                    if "TERMINADO" in est_str:
                        tag = 'terminado'
                    elif "BANEADA" in est_str:
                        tag = 'baneada'
                    else:
                        tag = 'abandonado'

                    if "BANEADA" in est_str:
                        wr_visual = "N/A"
                    else:
                        wr_visual = f"{wr}%" if wr is not None else "-"

                    self.tabla_historial.insert("", "end", values=(
                        contador_visual, 
                        booster, 
                        f"{user} -> {elo_fin}",     
                        wr_visual,
                        f_v(ini_raw), 
                        f_v(fin_raw), 
                        duracion, 
                        est_str    
                    ), tags=(tag,))
                    contador_visual += 1
            self.lbl_totales_h.configure(text=f"MOSTRANDO {len(self.tabla_historial.get_children())} REGISTROS")
        except Exception as e:
            print(f"Error en historial: {e}")

    # =========================================================================
    # SECCIÓN: REPORTES AVANZADOS Y GRÁFICOS
    # =========================================================================

    def mostrar_reportes(self):
        self.limpiar_pantalla()
        
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", background="#2b2b2b", foreground="white", fieldbackground="#2b2b2b", rowheight=30, borderwidth=0)
        style.configure("Treeview.Heading", background="#1a1a1a", foreground="white", relief="flat", font=("Arial", 10, "bold"))
        style.map("Treeview", background=[("selected", "#1f538d")], foreground=[("selected", "white")])

        main_container = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        main_container.pack(expand=True, fill="both", padx=30, pady=20)

        filtros_frame = ctk.CTkFrame(main_container, fg_color="#1a1a1a", corner_radius=15)
        filtros_frame.pack(fill="x", pady=(0, 20))
        ctk.CTkLabel(filtros_frame, text="📊 ANALÍTICA FINANCIERA", font=("Arial", 16, "bold")).pack(side="left", padx=20)
        
        self.combo_mes = ctk.CTkOptionMenu(filtros_frame, width=120,
            values=["Todos", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"])
        self.combo_mes.pack(side="left", padx=10)
        
        try:
            boosters = ["Todos"] + [b[1] for b in obtener_boosters_db()]
        except: boosters = ["Todos"]
        self.combo_booster_rep = ctk.CTkOptionMenu(filtros_frame, width=140, values=boosters)
        self.combo_booster_rep.pack(side="left", padx=10)

        ctk.CTkButton(filtros_frame, text="Calcular", width=100, command=self.actualizar_analitica).pack(side="left", padx=10)
        ctk.CTkButton(filtros_frame, text="📥 Exportar Excel", fg_color="#1d6f42", hover_color="#145231",
                      command=self.exportar_excel_avanzado).pack(side="right", padx=20)

        self.graficos_frame = ctk.CTkFrame(main_container, fg_color="#1a1a1a", height=280)
        self.graficos_frame.pack(fill="x", pady=10)
        self.graficos_frame.pack_propagate(False) 

        self.kpi_frame = ctk.CTkFrame(main_container, fg_color="transparent")
        self.kpi_frame.pack(fill="x", pady=10)

        table_frame = ctk.CTkFrame(main_container, fg_color="transparent")
        table_frame.pack(fill="both", expand=True, pady=10)

        cols = ("#", "booster", "elo", "demora", "pago_b", "ganancia_p", "bote", "total")
        self.tabla_rep = ttk.Treeview(table_frame, columns=cols, show="headings", height=8)
        self.tabla_rep.bind("<Button-3>", self.abrir_menu_contextual)
        self.tabla_rep.tag_configure('pagado', foreground='#2ecc71')
        self.tabla_rep.tag_configure('pendiente', foreground='#808080')
        
        headers = {
            "#": "N°", 
            "booster": "STAFF", 
            "elo": "ELO FINAL", 
            "demora": "DÍAS", 
            "pago_b": "PAGO STAFF", 
            "ganancia_p": "MI NETO", 
            "bote": "BOTE",           
            "total": "TOTAL CLI"
        }

        anchos = [40, 120, 100, 60, 100, 100, 80, 100]
        
        for c, w in zip(cols, anchos):
            self.tabla_rep.heading(c, text=headers[c])
            self.tabla_rep.column(c, width=w, anchor="center")

        scrollbar = ctk.CTkScrollbar(table_frame, orientation="vertical", command=self.tabla_rep.yview)
        self.tabla_rep.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.tabla_rep.pack(side="left", fill="both", expand=True)

        self.actualizar_analitica()

    def abrir_menu_contextual(self, event):
        item_id = self.tabla_rep.identify_row(event.y)
        if item_id:
            self.tabla_rep.selection_set(item_id)
            menu = tk.Menu(self, tearoff=0, bg="#1a1a1a", fg="white", activebackground="#1f538d", activeforeground="white")
            menu.add_command(label="📝 Editar Datos", command=self.ventana_edicion_rapida)
            menu.post(event.x_root, event.y_root)

    def ventana_edicion_rapida(self):
        seleccion = self.tabla_rep.selection()
        if not seleccion: return
        id_pedido = seleccion[0]
        lista_staff = []
        lista_elos = []

        try:
            conn = sqlite3.connect("perezboost.db")
            cur = conn.cursor()

            cur.execute("""
                SELECT booster_nombre, elo_final, pago_cliente, pago_booster, user_pass, wr, fecha_fin_real, pago_realizado, ganancia_empresa, bote_pedido, bote_wr
                FROM pedidos WHERE id = ?
            """, (id_pedido,))
            datos_db = cur.fetchone()

            try:
                cur.execute("SELECT nombre FROM boosters")
                lista_staff = [r[0] for r in cur.fetchall()]
            except: pass

            try:
                cur.execute("PRAGMA table_info(config_precios)")
                columnas = [c[1] for c in cur.fetchall()]
                col_elo = "division" if "division" in columnas else "elo"
                if col_elo in columnas:
                    cur.execute(f"SELECT {col_elo} FROM config_precios")
                    lista_elos = [r[0] for r in cur.fetchall()]
            except: pass

            conn.close()

        except Exception as e:
            messagebox.showerror("Error", f"Error leyendo BD: {e}")
            return

        if not datos_db: return

        staff_actual, elo_actual, pago_cli_actual, pago_staff_actual, user_pass, wr_actual, fecha_fin_actual, pago_actual, ganancia_actual, bote_ped, bote_wr = datos_db

        pago_cli_actual = float(pago_cli_actual or 0)
        pago_staff_actual = float(pago_staff_actual or 0)
        ganancia_actual = float(ganancia_actual or 0)
        bote_ped = float(bote_ped or 0)
        bote_wr = float(bote_wr or 0)

        v = ctk.CTkToplevel(self)
        v.title(f"Edición Pedido #{id_pedido}")
        v.geometry("380x830")
        v.attributes("-topmost", True)

        ctk.CTkLabel(v, text=f"Cuenta: {user_pass}", font=("Arial", 12, "bold"), text_color="#3498db").pack(pady=(15, 5))

        ctk.CTkLabel(v, text="Staff:", font=("Arial", 11, "bold")).pack(pady=(5,0))
        combo_staff = ctk.CTkOptionMenu(v, values=lista_staff, width=250)
        combo_staff.set(staff_actual); combo_staff.pack(pady=5)

        ctk.CTkLabel(v, text="Elo Final:", font=("Arial", 11, "bold")).pack(pady=(5,0))
        combo_elo = ctk.CTkOptionMenu(v, values=lista_elos, width=250)
        combo_elo.set(elo_actual if elo_actual else "Unranked"); combo_elo.pack(pady=5)

        ctk.CTkLabel(v, text="Fecha de Entrega (YYYY-MM-DD):", font=("Arial", 11, "bold")).pack(pady=(5, 0))
        e_fecha = ctk.CTkEntry(v, width=250)
        e_fecha.insert(0, str(fecha_fin_actual).split(" ")[0] if fecha_fin_actual else ""); e_fecha.pack(pady=2)

        ctk.CTkLabel(v, text="Win Rate (%):", font=("Arial", 11, "bold")).pack(pady=(5, 0))
        e_wr = ctk.CTkEntry(v, width=250)
        e_wr.insert(0, str(wr_actual if wr_actual is not None else "0")); e_wr.pack(pady=2)

        ctk.CTkLabel(v, text="Cobro al Cliente ($):", font=("Arial", 11, "bold")).pack(pady=(5, 0))
        e_cobro = ctk.CTkEntry(v, width=250)
        e_cobro.insert(0, str(pago_cli_actual)); e_cobro.pack(pady=2)

        ctk.CTkLabel(v, text="Pago al Staff ($):", font=("Arial", 11, "bold")).pack(pady=(5, 0))
        e_staff = ctk.CTkEntry(v, width=250)
        e_staff.insert(0, str(pago_staff_actual)); e_staff.pack(pady=2)

        ctk.CTkLabel(v, text="Bono Pedido aportado ($):", font=("Arial", 11, "bold"), text_color="#f1c40f").pack(pady=(5, 0))
        e_b_ped = ctk.CTkEntry(v, width=250)
        e_b_ped.insert(0, str(bote_ped)); e_b_ped.pack(pady=2)

        ctk.CTkLabel(v, text="Bono WR aportado ($):", font=("Arial", 11, "bold"), text_color="#f1c40f").pack(pady=(5, 0))
        e_b_wr = ctk.CTkEntry(v, width=250)
        e_b_wr.insert(0, str(bote_wr)); e_b_wr.pack(pady=2)

        ctk.CTkLabel(v, text="Mi Neto / Ganancia ($):", font=("Arial", 11, "bold"), text_color="#2ecc71").pack(pady=(5, 0))
        e_neto = ctk.CTkEntry(v, width=250, text_color="#2ecc71", font=("Arial", 12, "bold"))
        e_neto.insert(0, f"{ganancia_actual:.2f}"); e_neto.pack(pady=2)
        e_neto.configure(state="readonly")

        def recalcular_neto(event=None):
            try:
                c = float(e_cobro.get().replace("$","").strip() or 0)
                s = float(e_staff.get().replace("$","").strip() or 0)
                bp = float(e_b_ped.get().replace("$","").strip() or 0)
                bw = float(e_b_wr.get().replace("$","").strip() or 0)
                n = c - s - bp - bw
                e_neto.configure(state="normal")
                e_neto.delete(0, 'end')
                e_neto.insert(0, f"{n:.2f}")
                e_neto.configure(state="readonly")
            except: pass

        e_cobro.bind("<KeyRelease>", recalcular_neto)
        e_staff.bind("<KeyRelease>", recalcular_neto)
        e_b_ped.bind("<KeyRelease>", recalcular_neto)
        e_b_wr.bind("<KeyRelease>", recalcular_neto)

        ctk.CTkLabel(v, text="Estado del Pago:", font=("Arial", 11, "bold")).pack(pady=(10,0))
        var_pago = ctk.BooleanVar(value=True if pago_actual == 1 else False)
        switch_pago = ctk.CTkSwitch(v, text="Marcar como PAGADO", variable=var_pago, progress_color="#2ecc71")
        switch_pago.pack(pady=5)

        def guardar():
            try:
                nuevo_estado_pago = 1 if var_pago.get() else 0
                nueva_fecha = e_fecha.get().strip()
                if not nueva_fecha: nueva_fecha = None

                val_wr = float(e_wr.get().replace("%","").strip() or 0)
                val_cobro = float(e_cobro.get().replace("$","").strip() or 0)
                val_staff = float(e_staff.get().replace("$","").strip() or 0)
                val_bp = float(e_b_ped.get().replace("$","").strip() or 0)
                val_bw = float(e_b_wr.get().replace("$","").strip() or 0)
                val_neto = val_cobro - val_staff - val_bp - val_bw

                conn = sqlite3.connect("perezboost.db")
                cur = conn.cursor()

                cur.execute("""
                    UPDATE pedidos 
                    SET booster_nombre=?, elo_final=?, wr=?, pago_cliente=?, pago_booster=?, ganancia_empresa=?, fecha_fin_real=?, pago_realizado=?, bote_pedido=?, bote_wr=?
                    WHERE id=?
                """, (combo_staff.get(), combo_elo.get(), val_wr, val_cobro, val_staff, val_neto, nueva_fecha, nuevo_estado_pago, val_bp, val_bw, id_pedido))

                conn.commit(); conn.close()
                v.destroy()
                self.actualizar_analitica()
                messagebox.showinfo("Éxito", f"Registro actualizado.\nNeto Real: ${val_neto:.2f}")
            except Exception as e: messagebox.showerror("Error", f"Error al guardar: {e}")

        ctk.CTkButton(v, text="💾 Guardar Cambios", fg_color="#27ae60", hover_color="#2ecc71", height=40, width=200, command=guardar).pack(pady=(15, 10))

    def actualizar_analitica(self):
        
        for widget in self.kpi_frame.winfo_children(): widget.destroy()
        for i in self.tabla_rep.get_children(): self.tabla_rep.delete(i)
        for widget in self.graficos_frame.winfo_children(): widget.destroy()
        
        mes_sel = self.combo_mes.get()
        booster_sel = self.combo_booster_rep.get()

        datos = obtener_datos_reporte_avanzado(mes_sel, booster_sel)

        t_staff, t_neto, t_bote, t_ventas = 0.0, 0.0, 0.0, 0.0
        conteo_pagados, dias_totales = 0, 0

        def limpiar(v):
            try: return float(str(v).replace('$','').replace(',','').strip()) if v else 0.0
            except: return 0.0

        if not datos:
            ctk.CTkLabel(self.graficos_frame, text="⚠️ Sin registros finalizados").pack(expand=True)
            return

        contador_visual = 1

        for r in datos:
            try: esta_pagado = int(r[15] if r[15] is not None else 0)
            except: esta_pagado = 0
            
            v_total_cli = limpiar(r[11])
            v_pago_staff = limpiar(r[12])
            mi_neto_real = limpiar(r[13])

            txt_dias = "⚡ <24h"
            d_num = 0
            try:
                f_ini_str = str(r[5]).split(' ')[0] if r[5] else ""
                f_fin_str = str(r[10]).split(' ')[0] if r[10] else ""
                if f_ini_str and f_fin_str:
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                        try:
                            d_ini = datetime.strptime(f_ini_str, fmt)
                            d_fin = datetime.strptime(f_fin_str, fmt)
                            d_num = (d_fin - d_ini).days
                            txt_dias = f"{max(d_num, 1)} días"
                            break
                        except: continue
            except: txt_dias = "N/A"

            try: wr = float(r[9]) if r[9] else 0.0
            except: wr = 0.0
            
            b_ped, b_wr = self.obtener_bonos_actuales()
            valor_bote = (b_ped + b_wr) if wr >= 60 else b_ped
            valor_bote = v_total_cli - v_pago_staff - mi_neto_real

            if esta_pagado == 1:
                t_staff += v_pago_staff
                t_neto += mi_neto_real
                t_bote += valor_bote
                t_ventas += v_total_cli
                conteo_pagados += 1
                dias_totales += max(d_num, 1)
                tag_fila = 'pagado'
            else:
                tag_fila = 'pendiente'

            self.tabla_rep.insert("", "end", iid=r[0], values=(
                contador_visual, r[2], r[8], txt_dias,
                f"${v_pago_staff:.2f}", f"${mi_neto_real:.2f}",
                f"${valor_bote:.2f}", f"${v_total_cli:.2f}"
            ), tags=(tag_fila,))
            
            contador_visual += 1

        if mes_sel in ["Todos", "Enero"]:
            t_bote -= 5.0
            t_neto += 5.0

        prom_dias = dias_totales / conteo_pagados if conteo_pagados > 0 else 0

        self.crear_card_mini(self.kpi_frame, "COSTO STAFF", f"${t_staff:.2f}", "#3498db", 0)
        self.crear_card_mini(self.kpi_frame, "MI NETO REAL", f"${t_neto:.2f}", "#2ecc71", 1)
        self.crear_card_mini(self.kpi_frame, "BOTE RECOLECTADO", f"${t_bote:.2f}", "#f1c40f", 2)
        self.crear_card_mini(self.kpi_frame, "VENTAS TOTALES", f"${t_ventas:.2f}", "#9b59b6", 3)
        self.crear_card_mini(self.kpi_frame, "VELOCIDAD MEDIA", f"{prom_dias:.1f} d", "#e67e22", 4)
        self.dibujar_grafico_financiero(t_neto, t_staff, t_bote, booster_sel)

    def dibujar_grafico_financiero(self, total_perez, total_staff, total_bote, nombre_filtro):
        plt.close('all')
        plt.rcParams.update({
            'figure.facecolor': '#1a1a1a', 'axes.facecolor': '#1a1a1a',
            'axes.edgecolor': '#444444', 'axes.labelcolor': 'white',
            'xtick.color': 'white', 'ytick.color': 'white', 'text.color': 'white'
        })

        fig, ax = plt.subplots(figsize=(6, 2.8), dpi=100)
        
        label_izquierda = "Staff Total" if nombre_filtro == "Todos" else nombre_filtro

        categorias = [label_izquierda, 'Perez (Neto)', 'Bote Ranking']
        valores = [total_staff, total_perez, total_bote]
        colores = ['#3498db', '#2ecc71', '#f1c40f']

        barras = ax.bar(categorias, valores, color=colores, width=0.5, zorder=3)

        ax.set_title(f'DISTRIBUCIÓN FINANCIERA: {label_izquierda.upper()}', fontsize=11, fontweight='bold', pad=15)
        ax.set_ylabel('USD ($)', fontsize=9, fontweight='bold')
        ax.grid(axis='y', linestyle='--', alpha=0.3, zorder=0)

        for b in barras:
            height = b.get_height()
            if height > 0: 
                ax.annotate(f'${height:,.2f}', 
                            xy=(b.get_x() + b.get_width()/2, height),
                            xytext=(0, 5), textcoords="offset points", 
                            ha='center', va='bottom', weight='bold', fontsize=9, color="white")

        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        
        fig.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=self.graficos_frame)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.configure(bg='#1a1a1a', highlightthickness=0)
        canvas_widget.pack(fill="both", expand=True)
        canvas.draw()

    def crear_card_mini(self, master, titulo, valor, color, col):
        card = ctk.CTkFrame(master, fg_color="#1a1a1a", border_width=1, border_color=color)
        card.grid(row=0, column=col, padx=5, pady=5, sticky="nsew") 
        master.grid_columnconfigure(col, weight=1)
        ctk.CTkLabel(card, text=titulo, font=("Arial", 10, "bold"), text_color=color).pack(pady=(8,0))
        ctk.CTkLabel(card, text=valor, font=("Arial", 16, "bold"), text_color="white").pack(pady=(2,8))

    def exportar_excel_avanzado(self):
        import pandas as pd
        from datetime import datetime
        from tkinter import messagebox, filedialog
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        mes = self.combo_mes.get()
        booster = self.combo_booster_rep.get()
        nombre_sugerido = f"Cierre_Financiero_{mes}_{booster.replace(' ', '_')}.xlsx"
        
        datos = obtener_datos_reporte_avanzado(mes, booster)
        if not datos: 
            messagebox.showwarning("Sin Datos", f"No hay datos en {mes} para exportar.")
            return

        ruta = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=nombre_sugerido)
        if not ruta: return

        lista_procesada = []
        for r in datos:
            if str(r[7]).upper() == 'TERMINADO':
                try:
                    str_ini = str(r[5])[:10] if r[5] else ""
                    str_fin = str(r[10])[:10] if r[10] else ""
                    if str_ini and str_fin:
                        ini = pd.to_datetime(str_ini, format='mixed', dayfirst=False)
                        fin = pd.to_datetime(str_fin, format='mixed', dayfirst=False)
                        dias = (fin - ini).days
                        dias = max(dias, 1)
                    else: dias = 1
                except: dias = 1

                wr = float(r[9] if r[9] else 0)
                val_cliente = float(r[11] if r[11] else 0)
                val_staff = float(r[12] if r[12] else 0)
                val_perez_neto = float(r[13] if r[13] else 0)
                
                try: bp_e = float(r[18] if r[18] else 0)
                except: bp_e = 0.0
                try: bw_e = float(r[19] if r[19] else 0)
                except: bw_e = 0.0
                
                val_bote = (bp_e + bw_e) if (bp_e + bw_e) > 0 else (val_cliente - val_staff - val_perez_neto)
                
                lista_procesada.append({
                    "BOOSTER": r[2],
                    "CUENTA (User)": str(r[3]).split(':')[0] if r[3] else "N/A",
                    "ELO FINAL": r[8],
                    "WR (%)": wr / 100.0,
                    "DÍAS": dias,
                    "PAGO CLIENTE": val_cliente,
                    "PAGO STAFF": val_staff,
                    "APORTE BOTE": val_bote,
                    "GANANCIA NETA": val_perez_neto
                })

        if not lista_procesada: return
        
        df = pd.DataFrame(lista_procesada)
        totales = df.sum(numeric_only=True)
        row_total = pd.DataFrame([{
            "BOOSTER": "TOTALES ACUMULADOS >>", 
            "CUENTA (User)": "", "ELO FINAL": "", "WR (%)": "", "DÍAS": "",
            "PAGO CLIENTE": totales["PAGO CLIENTE"],
            "PAGO STAFF": totales["PAGO STAFF"],
            "APORTE BOTE": totales["APORTE BOTE"],
            "GANANCIA NETA": totales["GANANCIA NETA"]
        }])
        df = pd.concat([df, row_total], ignore_index=True) 

        try:
            with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Cierre_Financiero', index=False, startrow=1)
                ws = writer.sheets['Cierre_Financiero']

                borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                centrado = Alignment(horizontal='center', vertical='center')
                moneda = Alignment(horizontal='right', vertical='center')

                ws.merge_cells('A1:I1')
                titulo = ws['A1']
                titulo.value = f"REPORTE FINANCIERO PEREZBOOST | Mes: {mes.upper()} | Staff: {booster.upper()}"
                titulo.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
                titulo.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                titulo.alignment = centrado

                for col_num in range(1, len(df.columns) + 1):
                    celda = ws.cell(row=2, column=col_num)
                    celda.font = Font(bold=True, color="FFFFFF")
                    celda.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
                    celda.alignment = centrado
                    celda.border = borde_fino

                max_row = ws.max_row
                for row in range(3, max_row + 1):
                    is_total_row = (row == max_row)
                    for col in range(1, len(df.columns) + 1):
                        celda = ws.cell(row=row, column=col)
                        celda.border = borde_fino
                        if is_total_row:
                            celda.font = Font(bold=True)
                            celda.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        
                        if col == 4 and not is_total_row:
                            celda.number_format = '0.0%'
                            celda.alignment = centrado
                        elif col in [1, 2, 3, 5]:
                            celda.alignment = centrado
                        elif col >= 6:
                            celda.number_format = '"$" #,##0.00'
                            celda.alignment = moneda
                            if col == 9 and not is_total_row and celda.value and float(celda.value) > 0:
                                celda.font = Font(color="006100")
                                celda.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                for col_cells in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(col_cells[0].column)
                    for cell in col_cells:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column_letter].width = max_length + 4
                ws.freeze_panes = 'A3'

            messagebox.showinfo("Éxito", "Reporte generado correctamente.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar: {e}")
    

    # =========================================================================
    # SECCIÓN: LEADERBOARD
    # =========================================================================

    def mostrar_leaderboard(self):
        self.limpiar_pantalla()
        self.configurar_estilo_tabla()

        meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        fecha_hoy = datetime.now()
        mes_actual_idx = fecha_hoy.month - 1
        anio_actual = fecha_hoy.year

        main_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        main_frame.pack(expand=True, fill="both", padx=30, pady=20)

        header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        header_frame.pack(fill="x", pady=(0, 15))
        ctk.CTkLabel(header_frame, text="🏆 HALL OF FAME", font=("Arial", 28, "bold"), text_color="#ecf0f1").pack(side="left")

        self.combo_mes_rank = ctk.CTkOptionMenu(header_frame, values=meses_nombres, width=140, fg_color="#34495e")
        self.combo_mes_rank.set(meses_nombres[mes_actual_idx]) 
        self.combo_mes_rank.pack(side="right")
        ctk.CTkLabel(header_frame, text="📅 Filtrar Mes:", font=("Arial", 14)).pack(side="right", padx=10)

        prize_frame = ctk.CTkFrame(main_frame, fg_color="#2c3e50")
        prize_frame.pack(fill="x", pady=(0, 20), padx=5)
        self.lbl_bote = ctk.CTkLabel(prize_frame, text="...", font=("Arial", 18, "bold"), text_color="#ecf0f1")
        self.lbl_bote.pack(pady=15)

        stats_frame = ctk.CTkFrame(main_frame, fg_color="#1a1a1a", border_width=1, border_color="#5865F2")
        stats_frame.pack(fill="x", pady=(0, 20))
        stats_frame.columnconfigure((0,1,2), weight=1)

        ctk.CTkLabel(stats_frame, text="📦 Pedidos Totales", font=("Arial", 11), text_color="gray").grid(row=0, column=0, pady=(10,0))
        self.lbl_totales = ctk.CTkLabel(stats_frame, text="0", font=("Arial", 16, "bold"), text_color="white")
        self.lbl_totales.grid(row=1, column=0, pady=(0,10))

        ctk.CTkLabel(stats_frame, text="⚡ Eficiencia (Días/Pedido)", font=("Arial", 11), text_color="gray").grid(row=0, column=1, pady=(10,0))
        self.lbl_eficiencia = ctk.CTkLabel(stats_frame, text="0 Días", font=("Arial", 16, "bold"), text_color="#2ecc71")
        self.lbl_eficiencia.grid(row=1, column=1, pady=(0,10))

        ctk.CTkLabel(stats_frame, text="📊 WR Global", font=("Arial", 11), text_color="gray").grid(row=0, column=2, pady=(10,0))
        self.lbl_wr = ctk.CTkLabel(stats_frame, text="0%", font=("Arial", 16, "bold"), text_color="#f1c40f")
        self.lbl_wr.grid(row=1, column=2, pady=(0,10))

        tabla_frame = ctk.CTkFrame(main_frame, fg_color="#1e1e1e", corner_radius=10)
        tabla_frame.pack(expand=True, fill="both", pady=(0, 20))

        cols = ("Rango", "Staff", "Completados", "High_WR", "Abandonos", "Puntaje")
        self.tabla_rank = ttk.Treeview(tabla_frame, columns=cols, show="headings", height=8)

        headers = ["Rango", "Staff", "Terminados", "WR => 60%", "Abandonos", "Puntaje"]
        anchos = [80, 150, 100, 100, 80, 100]

        for col, title, ancho in zip(cols, headers, anchos):
            self.tabla_rank.heading(col, text=title)
            self.tabla_rank.column(col, width=ancho, anchor="center")
        
        self.tabla_rank.pack(side="left", expand=True, fill="both", padx=5, pady=5)
        self.tabla_rank.tag_configure("top1", background="#2d2d2d") 

        def actualizar_datos_ranking(event=None):
            nombre_mes = self.combo_mes_rank.get()
            num_mes = str(meses_nombres.index(nombre_mes) + 1).zfill(2)
            filtro = f"{anio_actual}-{num_mes}"

            for item in self.tabla_rank.get_children(): self.tabla_rank.delete(item)

            try:
                cant_term, cant_aban, wr_prom, cant_high_wr, avg_dias = obtener_resumen_mensual_db(filtro)

                ajuste_manual = 0.0
                nota_ajuste = ""

                if nombre_mes == "Febrero" and anio_actual == 2026:
                    ajuste_manual = 11.0
                    nota_ajuste = "($11 Enero Acumulado + $1 Pedidos + $1 WR)"
                elif nombre_mes == "Enero" and anio_actual == 2026:
                    ajuste_manual = -5.0
                    nota_ajuste = "($1 Pedido + $1 WR - $5 WR Pagados)"
                else:
                    nota_ajuste = "($1 Pedido + $1 Bono WR)"

                total_pedidos = cant_term
                b_ped, b_wr = self.obtener_bonos_actuales()
                bote_total = (float(cant_term) * b_ped) + (float(cant_high_wr) * b_wr) + ajuste_manual
                if bote_total < 0: bote_total = 0.0
                
                if nombre_mes == "Febrero" and anio_actual == 2026:
                    nota_ajuste = f"($11 Enero + ${b_ped:g} Pedidos + ${b_wr:g} WR)"
                elif nombre_mes == "Enero" and anio_actual == 2026:
                    nota_ajuste = f"(${b_ped:g} Pedido + ${b_wr:g} WR - $5 Pagados)"
                else:
                    nota_ajuste = f"(${b_ped:g} Pedido + ${b_wr:g} Bono WR)"

                if 0 < avg_dias < 1:
                    texto_eficiencia = "⚡ < 24 Horas"
                else:
                    texto_eficiencia = f"{avg_dias:.1f} Días"

            except Exception as e:
                print(f"Error calculo stats: {e}")
                cant_term, total_pedidos, avg_dias, wr_prom, cant_high_wr, bote_total, nota_ajuste = 0, 0, 0, 0, 0, 0, ""

            self.lbl_eficiencia.configure(text=texto_eficiencia)
            self.lbl_totales.configure(text=f"{cant_term}")
            self.lbl_wr.configure(text=f"{wr_prom:.1f}%")
            self.lbl_bote.configure(text=f"💰 BOTE {nombre_mes.upper()}: ${bote_total:.2f} USD 💰\n{nota_ajuste}")

            ranking_data = obtener_ranking_staff_db(filtro)
            
            if not ranking_data:
                self.tabla_rank.insert("", "end", values=("...", "Sin datos...", "-", "-", "-", "0 pts"))
            else:
                for i, b in enumerate(ranking_data, start=1):
                    nombre, terminados, aportes_wr, abandonos, puntaje = b[0], b[1], b[2], b[3], b[4]
                    rango = "🥇 MVP" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"#{i}"
                    item = self.tabla_rank.insert("", "end", values=(
                        rango, nombre, terminados, f"{aportes_wr}", abandonos, f"{int(puntaje)} pts"
                    ))
                    if i == 1: self.tabla_rank.item(item, tags=("top1",))

        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=10)
        
        ctk.CTkButton(btn_frame, text="🔄 Recargar", command=actualizar_datos_ranking, 
                    fg_color="#1f538d", height=40).pack(side="left", expand=True, padx=10)

        ctk.CTkButton(btn_frame, text="📢 Publicar Ranking", command=self.compartir_ranking_discord, 
                    fg_color="#5865F2", hover_color="#4752C4", height=40).pack(side="left", expand=True, padx=10)

        self.combo_mes_rank.configure(command=actualizar_datos_ranking)
        actualizar_datos_ranking()

    def compartir_ranking_discord(self):
        url = obtener_config_sistema("discord_webhook_ranking") 
        if not url: url = obtener_config_sistema("discord_webhook")
        
        if not url:
            messagebox.showerror("Error", "No hay Webhook configurado.")
            return

        mes_nombre = self.combo_mes_rank.get()
        meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        anio = datetime.now().year
        num_mes = str(meses_nombres.index(mes_nombre) + 1).zfill(2)
        filtro = f"{anio}-{num_mes}"

        try:
            cant_term, cant_aban, wr_prom, cant_high_wr, avg_dias = obtener_resumen_mensual_db(filtro)
            ajuste = 0.0
            msg_ajuste = ""
            if mes_nombre == "Febrero":
                ajuste = 11.0
                msg_ajuste = f"* 📈 Acumulado Enero: `$11.00 USD`\n"
            elif mes_nombre == "Enero":
                ajuste = -5.0
                msg_ajuste = f"* 💸 WR Pagado: `-$5.00 USD`\n"

            b_ped, b_wr = self.obtener_bonos_actuales()
            total_bote = (cant_term * b_ped) + (cant_high_wr * b_wr) + ajuste
        except:
            cant_term, cant_high_wr, total_bote, msg_ajuste = 0, 0, 0.0, ""

        ranking = obtener_ranking_staff_db(filtro)
        if not ranking:
            messagebox.showinfo("Vacío", f"No hay datos en {mes_nombre} para publicar.")
            return

        descripcion = f"# 🏆 HALL OF FAME - {mes_nombre.upper()} {anio}\n"
        descripcion += f"## 💰 BOTE ACUMULADO: `${total_bote:.2f} USD` 💰\n\n"
        descripcion += msg_ajuste
        descripcion += f"* ✅ Pedidos Terminados: `{cant_term}` (${b_ped:g} c/u)\n"
        descripcion += f"* 🔥 Bonos WR: `{cant_high_wr}` (${b_wr:g} c/u)\n"
        descripcion += "\n**DETALLE POR STAFF:**\n\n"
        
        for i, b in enumerate(ranking[:10], start=1):
            nombre, term, h_wr, aban, pts = b[0], b[1], b[2], b[3], b[4]
            icon = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"#{i}"
            descripcion += f"{icon} **{nombre}** — 🏆 `{int(pts)} Pts`\n"
            descripcion += f"   ✅ `{term}` Term.  |  🔥 `{h_wr}` Bonos WR  |  ❌ `{aban}` Drops\n\n"

        descripcion += "\nEl bote lo gana el primer lugar al final del mes."

        try:
            noti = DiscordNotifier(url)
            noti.enviar_notificacion(titulo="", descripcion=descripcion, color=0xFFD700, campos=[])
            messagebox.showinfo("Éxito", "Ranking publicado correctamente.")
        except Exception as e:
            messagebox.showerror("Error", f"Fallo al enviar: {e}")
            
    # =========================================================================
    # SECCIÓN: GESTIÓN FINANCIERA Y WALLET BINANCE
    # =========================================================================

    def asegurar_tabla_wallet(self):
        """Crea la tabla en la base de datos local si es la primera vez que se abre."""
        conn = conectar(); c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS wallet_perez (
                        id INTEGER PRIMARY KEY AUTOINCREMENT, 
                        fecha DATETIME DEFAULT CURRENT_TIMESTAMP, 
                        tipo TEXT, 
                        categoria TEXT, 
                        monto REAL, 
                        descripcion TEXT)''')
        conn.commit(); conn.close()

    def mostrar_finanzas(self):
        self.limpiar_pantalla()
        self.asegurar_tabla_wallet()
        self.configurar_estilo_tabla()
        
        self.main_scroll = ctk.CTkScrollableFrame(self.content_frame, fg_color="transparent")
        self.main_scroll.pack(fill="both", expand=True, padx=10, pady=10)

        header = ctk.CTkFrame(self.main_scroll, fg_color="transparent")
        header.pack(pady=(10, 15), padx=30, fill="x")
        ctk.CTkLabel(header, text="🏦 CAJA FUERTE Y BINANCE", font=("Arial", 24, "bold")).pack(side="left")

        # ---------------------------------------------------------
        # 1. CÁLCULO EN VIVO DE FONDOS (Historical Freeze)
        # ---------------------------------------------------------
        conn = conectar(); c = conn.cursor()
        c.execute("SELECT pago_cliente, pago_booster, ganancia_empresa FROM pedidos WHERE estado = 'Terminado' AND pago_realizado = 1")
        pedidos_pagados = c.fetchall()
        
        neto_hist = 0.0
        bote_hist = 0.0
        for p in pedidos_pagados:
            p_cli = float(p[0] or 0)
            p_boo = float(p[1] or 0)
            g_emp = float(p[2] or 0)
            
            bote_hist += (p_cli - p_boo - g_emp)
            neto_hist += g_emp

        neto_hist += 5.0
        bote_hist -= 5.0

        c.execute("SELECT tipo, categoria, monto FROM wallet_perez")
        movimientos = c.fetchall()
        for tipo, cat, monto in movimientos:
            m = float(monto or 0)
            if tipo == 'RETIRO':
                if cat == 'NETO': neto_hist -= m
                elif cat == 'BOTE': bote_hist -= m
            elif tipo == 'INGRESO':
                if cat == 'NETO': neto_hist += m
                elif cat == 'BOTE': bote_hist += m
        conn.close()

        total_binance = neto_hist + bote_hist

        # ---------------------------------------------------------
        # 2. TARJETAS DE SALDO
        # ---------------------------------------------------------
        wallet_frame = ctk.CTkFrame(self.main_scroll, fg_color="#111111", corner_radius=15, border_width=1, border_color="#333333")
        wallet_frame.pack(fill="x", padx=30, pady=5)
        
        cards_w = ctk.CTkFrame(wallet_frame, fg_color="transparent")
        cards_w.pack(fill="x", pady=20)
        cards_w.grid_columnconfigure((0,1,2), weight=1)

        def crear_card_binance(parent, titulo, valor, color, col):
            f = ctk.CTkFrame(parent, fg_color="#1a1a1a", border_width=1, border_color=color, corner_radius=10)
            f.grid(row=0, column=col, padx=15, sticky="nsew")
            ctk.CTkLabel(f, text=titulo, font=("Arial", 12, "bold"), text_color=color).pack(pady=(15,5))
            ctk.CTkLabel(f, text=f"${valor:,.2f}", font=("Arial", 30, "bold"), text_color="white").pack(pady=(0,15))

        crear_card_binance(cards_w, "MI NETO DISPONIBLE", neto_hist, "#2ecc71", 0)
        crear_card_binance(cards_w, "BOTE RANKING (STAFF)", bote_hist, "#f1c40f", 1)
        crear_card_binance(cards_w, "TOTAL BINANCE", total_binance, "#3498db", 2)

        # ---------------------------------------------------------
        # 3. SECCIÓN MEDIA: NUEVO MOVIMIENTO Y LIQUIDACIONES
        # ---------------------------------------------------------
        mid_frame = ctk.CTkFrame(self.main_scroll, fg_color="transparent")
        mid_frame.pack(fill="x", padx=30, pady=20)
        mid_frame.grid_columnconfigure(0, weight=1)
        mid_frame.grid_columnconfigure(1, weight=1)

        f_form = ctk.CTkFrame(mid_frame, fg_color="#1e1e1e", corner_radius=10)
        f_form.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        
        ctk.CTkLabel(f_form, text="💸 Retiros e Ingresos", font=("Arial", 16, "bold")).pack(pady=15)
        
        row1 = ctk.CTkFrame(f_form, fg_color="transparent"); row1.pack(fill="x", padx=20, pady=5)
        cb_tipo = ctk.CTkOptionMenu(row1, values=["RETIRO", "INGRESO"], width=140); cb_tipo.pack(side="left", padx=5)
        cb_cat = ctk.CTkOptionMenu(row1, values=["NETO", "BOTE"], width=140); cb_cat.pack(side="right", padx=5)
        
        row2 = ctk.CTkFrame(f_form, fg_color="transparent"); row2.pack(fill="x", padx=20, pady=5)
        e_monto = ctk.CTkEntry(row2, placeholder_text="Monto $", width=140); e_monto.pack(side="left", padx=5)
        e_desc = ctk.CTkEntry(row2, placeholder_text="Descripción (Ej: Nequi)", width=140); e_desc.pack(side="right", padx=5, expand=True, fill="x")

        def registrar_mov():
            try:
                m = float(e_monto.get().replace('$', '').strip())
                d = e_desc.get().strip()
                if not d: raise ValueError("Falta descripción")
                
                import time
                from datetime import datetime
                nuevo_id = int(time.time() * 1000) % 2147483647 
                fecha_local = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                conn = conectar(); cursor = conn.cursor()
                cursor.execute("INSERT INTO wallet_perez (id, fecha, tipo, categoria, monto, descripcion) VALUES (?,?,?,?,?,?)", 
                               (nuevo_id, fecha_local, cb_tipo.get(), cb_cat.get(), m, d))
                conn.commit(); conn.close()
                
                self.mostrar_finanzas()
                messagebox.showinfo("Éxito", "Transacción guardada exitosamente.")
            except Exception as e:
                messagebox.showerror("Error", "Monto inválido o falta descripción.")

        ctk.CTkButton(f_form, text="Registrar Transacción", fg_color="#3498db", command=registrar_mov).pack(pady=20)

        f_liq = ctk.CTkFrame(mid_frame, fg_color="#1e1e1e", corner_radius=10)
        f_liq.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        
        ctk.CTkLabel(f_liq, text="👤 Aprobar Pagos a Staff", font=("Arial", 16, "bold")).pack(pady=10)
        self.container_pagos = ctk.CTkScrollableFrame(f_liq, fg_color="transparent", height=150)
        self.container_pagos.pack(fill="both", expand=True, padx=10, pady=5)
        self.actualizar_lista_liquidaciones()

        # ---------------------------------------------------------
        # 4. HISTORIAL DE WALLET
        # ---------------------------------------------------------
        hist_frame = ctk.CTkFrame(self.main_scroll, fg_color="#1a1a1a", corner_radius=10)
        hist_frame.pack(fill="x", padx=30, pady=5)
        
        header_h = ctk.CTkFrame(hist_frame, fg_color="transparent")
        header_h.pack(fill="x", padx=15, pady=10)
        ctk.CTkLabel(header_h, text="📜 Historial de Wallet", font=("Arial", 14, "bold")).pack(side="left")
        ctk.CTkButton(header_h, text="🗑️ Eliminar Seleccionado", fg_color="#e74c3c", hover_color="#c0392b", width=150, height=25, command=self.eliminar_movimiento_wallet).pack(side="right")

        cols_w = ("id_v", "id_r", "fecha", "tipo", "cat", "monto", "desc")
        self.tabla_wallet = ttk.Treeview(hist_frame, columns=cols_w, show="headings", height=8)
        
        headers_w = ["Nº", "ID_R", "FECHA", "TIPO", "CAJA Afectada", "MONTO", "DESCRIPCIÓN"]
        widths_w = [40, 0, 150, 100, 100, 100, 400]
        
        for c, h, w in zip(cols_w, headers_w, widths_w):
            self.tabla_wallet.heading(c, text=h)
            if c == "id_r":
                self.tabla_wallet.column(c, width=0, stretch=tk.NO)
            else:
                self.tabla_wallet.column(c, width=w, anchor="center" if c != "desc" else "w")
            
        self.tabla_wallet.pack(fill="x", padx=15, pady=10)
        
        conn = conectar(); c = conn.cursor()
        c.execute("SELECT id, fecha, tipo, categoria, monto, descripcion FROM wallet_perez ORDER BY fecha DESC, id DESC")
        
        for i, row in enumerate(c.fetchall(), start=1):
            signo = "-" if row[2] == "RETIRO" else "+"
            monto_str = f"{signo}${row[4]:.2f}"
            self.tabla_wallet.insert("", "end", values=(i, row[0], str(row[1])[:16], row[2], row[3], monto_str, row[5]))
        conn.close()

    def actualizar_lista_liquidaciones(self):
        for widget in self.container_pagos.winfo_children():
            widget.destroy()
            
        saldos = obtener_saldos_pendientes_db()
        
        if not saldos:
            ctk.CTkLabel(self.container_pagos, text="✨ No hay deudas pendientes.", font=("Arial", 13), text_color="gray").pack(pady=40)
            return

        for booster, total, cant, detalle in saldos:
            card = ctk.CTkFrame(self.container_pagos, fg_color="#2b2b2b", height=70)
            card.pack(fill="x", pady=5)
            card.pack_propagate(False) 

            info = ctk.CTkFrame(card, fg_color="transparent")
            info.pack(side="left", padx=15, fill="x", expand=True)
            
            ctk.CTkLabel(info, text=f"{booster}  •  {cant} pedidos", font=("Arial", 13, "bold"), anchor="w").pack(anchor="w")
            
            if detalle and len(detalle) > 40: 
                detalle = detalle[:40] + "..." 
            
            ctk.CTkLabel(info, text=f"📄 {detalle}", font=("Arial", 10), text_color="#aaaaaa", anchor="w").pack(anchor="w")

            ctk.CTkButton(card, text=f"Pagar ${total:,.2f}", width=100, height=30,
                          fg_color="#2ecc71", hover_color="#27ae60", font=("Arial", 11, "bold"),
                          command=lambda b=booster, t=total: self.ejecutar_pago(b, t)).pack(side="right", padx=15)

    def ejecutar_pago(self, booster, total):
        if messagebox.askyesno("Confirmar Liquidación", f"¿Confirmas el pago de ${total:,.2f} a {booster}?\n\nEsta acción marcará sus pedidos como PAGADOS y se reflejará en Binance."):
            cant = liquidar_pagos_booster_db(booster)
            registrar_log("PAGO_STAFF", f"Liquidación de ${total:,.2f} a {booster} ({cant} pedidos)")
            messagebox.showinfo("Pago Exitoso", f"Se han liquidado {cant} pedidos de {booster} correctamente.")
            self.mostrar_finanzas()

    def eliminar_movimiento_wallet(self):
        sel = self.tabla_wallet.selection()
        if not sel: return
        id_mov = self.tabla_wallet.item(sel)['values'][1] 
        
        if messagebox.askyesno("Confirmar", "⚠️ ¿Eliminar este movimiento?\n\nSe recalcularán los saldos de Binance."):
            conn = conectar(); c = conn.cursor()
            c.execute("DELETE FROM wallet_perez WHERE id = ?", (id_mov,))
            conn.commit(); conn.close()
            self.mostrar_finanzas()

    # =========================================================================
    # SECCIÓN: POP-UPS (VENTANAS EMERGENTES)
    # =========================================================================

    def abrir_ventana_booster(self):
        v = ctk.CTkToplevel(self); v.title("Nuevo Staff"); self.centrar_ventana(v, 400, 300); v.attributes("-topmost", True)
        entry = ctk.CTkEntry(v, width=250); entry.pack(pady=30)
        def save():
            if entry.get() and agregar_booster(entry.get().strip().title()):
                v.destroy(); self.mostrar_boosters()
        ctk.CTkButton(v, text="Guardar", command=save, fg_color="#2ecc71").pack()

    def abrir_ventana_editar_booster(self):
        sel = self.tabla_boosters.selection()
        if not sel: return
        id_r = self.tabla_boosters.item(sel)['values'][1]
        v = ctk.CTkToplevel(self); self.centrar_ventana(v, 350, 250); v.attributes("-topmost", True)
        entry = ctk.CTkEntry(v, width=200); entry.pack(pady=20)
        def save():
            actualizar_booster_db(id_r, entry.get().strip().title()); v.destroy(); self.mostrar_boosters()
        ctk.CTkButton(v, text="Actualizar", fg_color="#27ae60", command=save).pack()

    def eliminar_booster_seleccionado(self):
        sel = self.tabla_boosters.selection()
        if sel and messagebox.askyesno("Confirmar", "¿Eliminar staff?", parent=self):
            id_r = self.tabla_boosters.item(sel)['values'][1]
            if eliminar_booster(id_r): self.mostrar_boosters()

    def abrir_ventana_editar_pedido(self):
        sel = self.tabla_pedidos.selection()
        if not sel: return

        val = self.tabla_pedidos.item(sel)['values']
        id_r = val[1]
        booster_actual = val[2]
        
        v = ctk.CTkToplevel(self)
        self.centrar_ventana(v, 400, 580)
        v.attributes("-topmost", True)
        v.title(f"Editar Pedido #{id_r}")
        
        entradas = {}

        import sqlite3
        try:
            conn = sqlite3.connect("perezboost.db")
            cur = conn.cursor()
            cur.execute("SELECT nombre FROM boosters ORDER BY nombre ASC")
            nombres_staff = [row[0] for row in cur.fetchall()]
            conn.close()
        except Exception as e:
            nombres_staff = [booster_actual]
            
        if not nombres_staff:
            nombres_staff = ["Sin Staff"]

        if booster_actual not in nombres_staff and booster_actual:
            nombres_staff.insert(0, booster_actual)

        ctk.CTkLabel(v, text="Staff:", font=("Arial", 12, "bold")).pack(pady=(10, 0))
        combo_staff = ctk.CTkComboBox(v, values=nombres_staff, width=280)
        combo_staff.set(booster_actual) 
        combo_staff.pack(pady=5)
        entradas["booster_nombre"] = combo_staff

        campos = [
            ("Elo Inicial:", "elo_inicial", val[3]),
            ("Cuenta / User:", "user_pass", val[4]), 
            ("Fecha Inicio (D/M/Y):", "fecha_inicio", val[5]),
            ("Fecha Entrega (D/M/Y):", "fecha_limite", val[6])
        ]
        
        for lab, col, act in campos:
            ctk.CTkLabel(v, text=lab, font=("Arial", 12, "bold")).pack(pady=(10, 0))
            e = ctk.CTkEntry(v, width=280)
            e.insert(0, act if act else "")
            e.pack(pady=5)
            entradas[col] = e

        def save():
            try:
                datos_nuevos = {}
                for k, e in entradas.items():
                    valor = e.get().strip()

                    if k in ["fecha_inicio", "fecha_limite"]:
                        if "/" in valor:
                            valor = datetime.strptime(valor, "%d/%m/%Y").strftime("%Y-%m-%d")
                        elif "-" in valor and len(valor) >= 10:
                            valor = valor.split(' ')[0]
                    
                    datos_nuevos[k] = valor

                actualizar_pedido_db(id_r, datos_nuevos)
                
                registrar_log("EDICION_PEDIDO", f"Pedido #{id_r} editado manualmente. Staff: {datos_nuevos['booster_nombre']}")
                v.destroy()
                self.mostrar_pedidos()
                
            except Exception as ex:
                messagebox.showerror("Error", f"Error al guardar:\n{ex}", parent=v)

        ctk.CTkButton(v, text="💾 Guardar Cambios", fg_color="#27ae60", hover_color="#1e8449",
                       height=40, font=("Arial", 12, "bold"), command=save).pack(pady=25)

    def abrir_ventana_finalizar(self):
        sel = self.tabla_pedidos.selection()
        if not sel: return

        val_fila = self.tabla_pedidos.item(sel)['values']
        id_r = val_fila[1]
        nom_booster = val_fila[2]

        try:
            conn = conectar()
            cursor = conn.cursor()
            mes_actual_iso = datetime.now().strftime("%Y-%m")
            cursor.execute("SELECT COUNT(*) FROM pedidos WHERE estado = 'Terminado' AND fecha_fin_real LIKE ?", (f"{mes_actual_iso}%",))
            num_orden_mes = cursor.fetchone()[0] + 1
            conn.close()
        except:
            num_orden_mes = "?"

        tarifas_raw = obtener_config_precios()
        tarifas = [t[0] for t in tarifas_raw]

        v = ctk.CTkToplevel(self)
        self.centrar_ventana(v, 400, 650)
        v.attributes("-topmost", True)
        v.title(f"Finalizar Orden #{num_orden_mes}")

        ctk.CTkLabel(v, text="¿En qué Elo quedó la cuenta?").pack(pady=(10,0))
        cb_div = ctk.CTkOptionMenu(v, values=tarifas, width=250)
        cb_div.pack(pady=5)

        ctk.CTkLabel(v, text="WinRate Final (%):").pack()
        e_wr = ctk.CTkEntry(v, placeholder_text="Ej: 85")
        e_wr.pack(pady=5)

        ctk.CTkLabel(v, text="Ajuste de Saldo ($):").pack(pady=(10,0))
        e_ajuste = ctk.CTkEntry(v, placeholder_text="0.00")
        e_ajuste.insert(0, "0.00")
        e_ajuste.pack(pady=5)

        def limpiar_ajuste(event):
            if e_ajuste.get() == "0.00":
                e_ajuste.delete(0, 'end')
        e_ajuste.bind("<Button-1>", limpiar_ajuste)

        # 🛑 SECCIÓN MANUAL DE BONOS
        try: bp_val = float(obtener_config_sistema("bono_pedido") or 1.0)
        except: bp_val = 1.0
        try: bwr_val = float(obtener_config_sistema("bono_wr") or 1.0)
        except: bwr_val = 1.0

        ctk.CTkLabel(v, text="💰 Aportes al Bote (Ranking):", font=("Arial", 12, "bold"), text_color="#f1c40f").pack(pady=(15,5))

        frame_bonos = ctk.CTkFrame(v, fg_color="transparent")
        frame_bonos.pack(pady=5)

        var_ped = ctk.BooleanVar(value=True) # Siempre True por defecto
        chk_ped = ctk.CTkCheckBox(frame_bonos, text=f"Bono Pedido (${bp_val})", variable=var_ped)
        chk_ped.pack(side="left", padx=10)

        var_wr = ctk.BooleanVar(value=False)
        chk_wr = ctk.CTkCheckBox(frame_bonos, text=f"Bono WR (${bwr_val})", variable=var_wr)
        chk_wr.pack(side="left", padx=10)

        var_publicar = ctk.BooleanVar(value=True) 
        chk_discord = ctk.CTkCheckBox(
            v, text="📢 Publicar resultado en Discord", 
            variable=var_publicar, fg_color="#5865F2",
            checkbox_height=20, checkbox_width=20
        )
        chk_discord.pack(pady=(20, 10))

        def finish():
            try:
                val_wr = e_wr.get()
                val_ajuste = e_ajuste.get()

                if not val_wr:
                    messagebox.showerror("Error", "El WinRate está vacío.", parent=v)
                    return

                wr = float(val_wr)
                ajuste = float(val_ajuste) if val_ajuste and val_ajuste.strip() != "" else 0.0

                elo_fin = cb_div.get()
                fecha_hoy_iso = datetime.now().strftime("%Y-%m-%d")

                conn = conectar(); cursor = conn.cursor()
                cursor.execute("SELECT precio_cliente, margen_perez FROM config_precios WHERE division = ?", (elo_fin,))
                tarifa = cursor.fetchone()

                aporte_ped = bp_val if var_ped.get() else 0.0
                aporte_wr = bwr_val if var_wr.get() else 0.0

                if tarifa:
                    p_cli_base = float(tarifa[0])
                    g_per_base = float(tarifa[1])

                    p_cliente = p_cli_base + aporte_wr
                    # Ningún trato especial para PEREZ
                    p_booster = (p_cli_base - g_per_base) + ajuste
                    g_perez = (g_per_base - aporte_ped) - ajuste
                else:
                    p_cliente, g_perez, p_booster = 0.0, 0.0, 0.0

                f_obj = datetime.now()
                nombres_meses = {1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO", 
                                7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"}
                nombre_mes_es = nombres_meses[f_obj.month]
                mes_cierre_iso = f_obj.strftime("%Y-%m")

                if finalizar_pedido_db(id_r, wr, fecha_hoy_iso, elo_fin, g_perez, p_booster, p_cliente, ajuste, aporte_ped, aporte_wr):
                    if var_publicar.get():
                        try:
                            cursor.execute("SELECT COUNT(*) FROM pedidos WHERE estado = 'Terminado' AND fecha_fin_real LIKE ?", (f"{mes_cierre_iso}%",))
                            num_orden_discord = cursor.fetchone()[0]
                            url = obtener_config_sistema("discord_webhook")

                            if url:
                                cursor.execute("SELECT COUNT(*) FROM pedidos WHERE booster_nombre = ? AND estado = 'Terminado' AND fecha_fin_real LIKE ?", (nom_booster, f"{mes_cierre_iso}%"))
                                total_mes_booster = cursor.fetchone()[0]
                                noti = DiscordNotifier(url)

                                campos_embed = [
                                    {"name": "👤 STAFF", "value": f"**{nom_booster}**", "inline": True},
                                    {"name": "🔥 RACHA MENSUAL", "value": f"**{total_mes_booster}º del mes**", "inline": True},
                                    {"name": "📌 QUEDÓ EN", "value": f"`{elo_fin}`", "inline": True},
                                    {"name": "🎯 WR", "value": f"`{wr}%`", "inline": True},
                                    {"name": "💸 PAGO", "value": f"**${p_booster:.2f}**", "inline": True},
                                    {"name": "⚙️ AJUSTE", "value": f"**${ajuste:.2f}**", "inline": True}
                                ]

                                noti.enviar_notificacion(
                                    titulo=f"✅ PEDIDO #{num_orden_discord} DE {nombre_mes_es}", 
                                    descripcion="", 
                                    color=5763719, 
                                    campos=campos_embed
                                )
                        except Exception as e:
                            print(f"Error notificando Discord: {e}")

                    conn.close() 
                    registrar_log("PEDIDO_FINALIZADO", f"Orden #{id_r} cerrada. Ajuste: {ajuste}.")
                    messagebox.showinfo("Éxito", f"¡Pedido finalizado!\nPago Staff: ${p_booster:.2f}\nTu Neto: ${g_perez:.2f}\nAporte al Bote: ${(aporte_ped + aporte_wr):.2f}", parent=v)
                    v.destroy()
                    self.mostrar_pedidos()
                else:
                    conn.close()
                    messagebox.showerror("Error", "No se pudo actualizar la DB.", parent=v)
            except ValueError:
                messagebox.showerror("Error", "WinRate y Ajuste deben ser números.", parent=v)

        ctk.CTkButton(v, text="Confirmar Finalización", fg_color="#2ecc71", height=40, font=("Arial", 12, "bold"), command=finish).pack(pady=20)
        
    def abrir_ventana_extender_tiempo(self):
        sel = self.tabla_pedidos.selection()
        if not sel: return
        id_r = self.tabla_pedidos.item(sel)['values'][1]
        v = ctk.CTkToplevel(self); self.centrar_ventana(v, 300, 200); v.attributes("-topmost", True)
        e_dias = ctk.CTkEntry(v, width=100); e_dias.insert(0, "1"); e_dias.pack(pady=20)
        def confirm():
            f_act = str(self.tabla_pedidos.item(sel)['values'][6])
            fecha_dt = None
            for formato in ("%Y-%m-%d", "%d/%m/%Y"):
                try:
                    fecha_dt = datetime.strptime(f_act, formato)
                    break
                except ValueError:
                    continue
            if fecha_dt:
                nueva = (fecha_dt + timedelta(days=int(e_dias.get()))).strftime("%Y-%m-%d")
                actualizar_pedido_db(id_r, {"fecha_limite": nueva})
                v.destroy()
                self.mostrar_pedidos()
            else:
                print(f"Error: No se pudo reconocer el formato de fecha: {f_act}")
        ctk.CTkButton(v, text="Extender", command=confirm).pack()

    def abrir_ventana_nuevo_pedido(self):
        
        b_raw = obtener_boosters_db(); elos = obtener_elos_en_stock()
        
        if not b_raw:
            messagebox.showwarning("Falta Staff", "No tienes Boosters registrados.")
            return

        if not elos:
            messagebox.showwarning("Sin Stock", "⚠️ No hay cuentas disponibles en el Inventario.")
            return
        
        map_b = {b[1]: b[0] for b in b_raw}
        v = ctk.CTkToplevel(self); self.centrar_ventana(v, 450, 650); v.attributes("-topmost", True)
        
        ctk.CTkLabel(v, text="NUEVO PEDIDO", font=("Arial", 16, "bold")).pack(pady=15)

        cb_b = ctk.CTkOptionMenu(v, values=list(map_b.keys()), width=300); cb_b.pack(pady=10)
        self.map_c_id = {}; self.map_c_note = {}
        
        def update_note(choice):
            n = self.map_c_note.get(choice, "FRESH")
            e_n.configure(state="normal"); e_n.delete(0, "end"); e_n.insert(0, n); e_n.configure(state="readonly")
        
        def change_elo(choice):
            data = obtener_cuentas_filtradas_datos(choice)
            if data:
                self.map_c_id = {c[1]: c[0] for c in data}; self.map_c_note = {c[1]: (c[2] if c[2] else "FRESH") for c in data}
                names = list(self.map_c_id.keys()); cb_c.configure(values=names); cb_c.set(names[0]); update_note(names[0])
        
        ctk.CTkLabel(v, text="Elo Inicial:").pack()
        cb_e = ctk.CTkOptionMenu(v, values=elos, width=300, command=change_elo); cb_e.pack(pady=5)
        
        ctk.CTkLabel(v, text="Cuenta a Asignar:").pack()
        cb_c = ctk.CTkOptionMenu(v, values=[], width=300, command=update_note); cb_c.pack(pady=5)
        
        ctk.CTkLabel(v, text="Nota de la Cuenta:").pack()
        e_n = ctk.CTkEntry(v, width=300, state="readonly"); e_n.pack(pady=5)
        
        ctk.CTkLabel(v, text="Días para entregar:").pack()
        e_d = ctk.CTkEntry(v, width=300); e_d.insert(0, "10"); e_d.pack(pady=5)
        
        if elos: change_elo(elos[0])
        
        def go():
            try:
                dias = int(e_d.get())
                f_fin = calcular_fecha_limite_sugerida(dias).split(' ')[0]
                b_name = cb_b.get()
                c_name = cb_c.get()
                if crear_pedido(map_b[cb_b.get()], cb_b.get(), self.map_c_id[cb_c.get()], cb_c.get(), cb_e.get(), f_fin):
                    registrar_log("NUEVO_PEDIDO", f"Asignado a {b_name}. Cuenta: {c_name}")
                    v.destroy(); self.mostrar_pedidos()
            except ValueError:
                messagebox.showerror("Error", "Los días deben ser un número entero.", parent=v)

        ctk.CTkButton(v, text="Iniciar Orden", fg_color="#2ecc71", command=go).pack(pady=20)

    def abrir_ventana_reportar_abandono(self):
        sel = self.tabla_pedidos.selection()
        if not sel: return
        id_r = self.tabla_pedidos.item(sel)['values'][1]
        cuenta = self.tabla_pedidos.item(sel)['values'][4] 
        
        v = ctk.CTkToplevel(self)
        self.centrar_ventana(v, 350, 450)
        v.attributes("-topmost", True)
        
        ctk.CTkLabel(v, text="🚫 REPORTAR DROP", font=("Arial", 16, "bold"), text_color="#e74c3c").pack(pady=(20, 10))
        
        e_elo = ctk.CTkEntry(v, placeholder_text="Elo dejado...", width=250)
        e_elo.pack(pady=10)
        
        e_wr = ctk.CTkEntry(v, placeholder_text="WR dejado (Ej: 45)...", width=250)
        e_wr.pack(pady=10)
        
        e_notas = ctk.CTkEntry(v, placeholder_text="Motivo (Ej: Troll en promo)...", width=250)
        e_notas.pack(pady=10)
        
        def confirm():
            elo_dejado = e_elo.get().upper().strip()
            wr_dejado = e_wr.get().strip()
            nota_texto = e_notas.get().strip()
            
            if registrar_abandono_db(id_r, elo_dejado, wr_dejado):
                registrar_log("ABANDONO_PEDIDO", f"Pedido #{id_r} Drop. Elo: {elo_dejado} | WR: {wr_dejado} | Razón: {nota_texto}")
                try:
                    detalles = []
                    if elo_dejado: detalles.append(elo_dejado)
                    if wr_dejado: detalles.append(f"{wr_dejado}% WR")
                    if nota_texto: detalles.append(nota_texto)
                    nota_final = "ABANDONADA"
                    if detalles:
                        nota_final += ": " + ", ".join(detalles)
                    
                    conn = conectar()
                    cursor = conn.cursor()
                    cursor.execute("UPDATE inventario SET descripcion = ? WHERE user_pass = ?", (nota_final, cuenta.strip()))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    print(f"Error guardando nota en inventario: {e}")
                v.destroy()
                self.mostrar_pedidos()
                messagebox.showinfo("Éxito", "Cuenta reportada como DROP. Se ha enviado al inventario y marcado como Abandonada.")
            else:
                messagebox.showerror("Error", "Ocurrió un problema interno al intentar abandonar la cuenta.")
            
            if registrar_abandono_db(id_r, elo_dejado, wr_dejado):
                registrar_log("ABANDONO_PEDIDO", f"Pedido #{id_r} Drop. Elo: {elo_dejado} | WR: {wr_dejado} | Razón: {nota_texto}")
                try:
                    detalles = []
                    if elo_dejado: detalles.append(elo_dejado)
                    if wr_dejado: detalles.append(f"{wr_dejado}% WR")
                    if nota_texto: detalles.append(nota_texto)
                    nota_final = "ABANDONADA"
                    if detalles:
                        nota_final += ": " + ", ".join(detalles)
                    
                    conn = conectar()
                    cursor = conn.cursor()
                    cursor.execute("UPDATE inventario SET descripcion = ? WHERE user_pass = ?", (nota_final, cuenta.strip()))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    print(f"Error guardando nota en inventario: {e}")
                v.destroy()
                self.mostrar_pedidos()
                
        ctk.CTkButton(v, text="Confirmar DROP", fg_color="#e74c3c", hover_color="#c0392b", height=40, command=confirm).pack(pady=20)
        
    def reportar_ban_seleccionado(self):
        sel = self.tabla_pedidos.selection()
        if not sel: return
        
        val_fila = self.tabla_pedidos.item(sel)['values']
        id_r = val_fila[1]
        cuenta = val_fila[4]

        if messagebox.askyesno("🔨 Confirmar Ban", f"¿Estás seguro de marcar la cuenta '{cuenta}' como BANEADA?\n\nDesaparecerá de activos y quedará en el historial.", parent=self):
            hoy_str = datetime.now().strftime("%Y-%m-%d %H:%M")
            actualizar_pedido_db(id_r, {
                "estado": "Baneada",
                "fecha_fin_real": hoy_str,
                "elo_final": "BANEADA",
                "wr": None
            })
            
            registrar_log("CUENTA_BANEADA", f"El pedido #{id_r} ({cuenta}) fue aniquilado (Ban).")
            self.mostrar_pedidos()

    def abrir_ventana_registro(self):
        v = ctk.CTkToplevel(self); self.centrar_ventana(v, 400, 450); v.attributes("-topmost", True)
        ctk.CTkLabel(v, text="AÑADIR CUENTA", font=("Arial", 16, "bold")).pack(pady=20)
        up = ctk.CTkEntry(v, placeholder_text="User:Pass", width=250); up.pack(pady=10)
        elo = ctk.CTkOptionMenu(v, values=["Emerald/Plat", "DIAMANTE"], width=250); elo.pack(pady=10)
        not_ent = ctk.CTkEntry(v, placeholder_text="Notas", width=250); not_ent.pack(pady=10)
        def save():
            if registrar_cuenta_gui(up.get(), elo.get(), not_ent.get())[0]:
                v.destroy(); self.mostrar_inventario()
        ctk.CTkButton(v, text="Guardar", command=save, fg_color="#2ecc71").pack(pady=20)

    def abrir_ventana_editar_inventario(self):
        sel = self.tabla_inv.selection()
        if not sel: return
        v_id = self.tabla_inv.item(sel)['values'][0]
        datos_r = next(d for d in self.datos_inventario if d[0] == v_id)
        id_real = datos_r[1]
        v = ctk.CTkToplevel(self); self.centrar_ventana(v, 400, 450); v.attributes("-topmost", True)
        ctk.CTkLabel(v, text="EDITAR CUENTA", font=("Arial", 16, "bold")).pack(pady=20)
        e_up = ctk.CTkEntry(v, width=250); e_up.insert(0, datos_r[2]); e_up.pack(pady=10)
        e_elo = ctk.CTkOptionMenu(v, values=["Emerald/Plat", "DIAMANTE"], width=250); e_elo.set(datos_r[3]); e_elo.pack(pady=10)
        e_not = ctk.CTkEntry(v, width=250); e_not.insert(0, datos_r[4]); e_not.pack(pady=10)
        def save():
            actualizar_inventario_db(id_real, {"user_pass": e_up.get(), "elo_tipo": e_elo.get(), "descripcion": e_not.get()})
            v.destroy(); self.mostrar_inventario()
        ctk.CTkButton(v, text="Actualizar", fg_color="#27ae60", command=save).pack(pady=20)

    def abrir_ventana_masivo(self):
        v = ctk.CTkToplevel(self); self.centrar_ventana(v, 500, 550); v.attributes("-topmost", True)
        txt = ctk.CTkTextbox(v, width=450, height=250); txt.pack(pady=10)
        elo = ctk.CTkOptionMenu(v, values=["Emerald/Plat", "DIAMANTE"], width=200); elo.pack(pady=10)
        def proc():
            registrar_lote_gui(txt.get("1.0", "end"), elo.get())
            v.destroy(); self.mostrar_inventario()
        ctk.CTkButton(v, text="🚀 Importar", command=proc, fg_color="#3498db").pack(pady=20)

    def eliminar_seleccionado(self):
        sel = self.tabla_inv.selection()
        if sel and messagebox.askyesno("Confirmar", "¿Eliminar cuenta permanentemente?", parent=self):
            v_id = self.tabla_inv.item(sel)['values'][0]
            cuenta_afectada = self.tabla_inv.item(sel)['values'][1]
            id_r = next(d[1] for d in self.datos_inventario if d[0] == v_id)
            if eliminar_cuenta_gui(id_r):
                registrar_log("STOCK_ELIMINADO", f"Cuenta eliminada: {cuenta_afectada}")
                self.mostrar_inventario()

    def abrir_ventana_nuevo_precio(self):
        v = ctk.CTkToplevel(self); self.centrar_ventana(v, 350, 480); v.attributes("-topmost", True)
        ctk.CTkLabel(v, text="NUEVA TARIFA", font=("Arial", 16, "bold")).pack(pady=20)
        ctk.CTkLabel(v, text="Puntos de Ranking:", font=("Arial", 11)).pack(pady=(5,0))
        
        e_div = ctk.CTkEntry(v, placeholder_text="Ej: D1", width=200); e_div.pack(pady=5)
        e_cli = ctk.CTkEntry(v, placeholder_text="Precio Cliente $", width=200); e_cli.pack(pady=5)
        e_per = ctk.CTkEntry(v, placeholder_text="Margen Perez $", width=200); e_per.pack(pady=5)
        e_pts = ctk.CTkEntry(v, placeholder_text="Ej: 45", width=200); e_pts.insert(0, "2"); e_pts.pack(pady=5)
        
        def save():
            try:
                div = e_div.get().upper()
                cli = float(e_cli.get())
                per = float(e_per.get())
                pts = int(e_pts.get())
                
                if agregar_precio_db(div, cli, per, pts):
                    v.destroy(); self.actualizar_tabla_precios()
            except ValueError:
                messagebox.showerror("Error", "Asegúrate de ingresar números válidos en Precio y Puntos.")

        ctk.CTkButton(v, text="Guardar", command=save, fg_color="#2ecc71").pack(pady=20)
    
    def abrir_ventana_editar_precio(self):
        sel = self.tabla_precios.selection()
        if not sel: return
        datos = self.tabla_precios.item(sel)['values']
        
        div = datos[0]
        p_cli = str(datos[1]).replace('$','').replace(',','')
        m_per = str(datos[2]).replace('$','').replace(',','')
        p_pts = str(datos[4])
        
        v = ctk.CTkToplevel(self); self.centrar_ventana(v, 350, 480); v.attributes("-topmost", True)
        ctk.CTkLabel(v, text=f"EDITAR {div}", font=("Arial", 16, "bold")).pack(pady=20)
        
        ctk.CTkLabel(v, text="Precio Cliente:").pack()
        e_cli = ctk.CTkEntry(v, width=200); e_cli.insert(0, p_cli); e_cli.pack(pady=5)
        
        ctk.CTkLabel(v, text="Margen Perez:").pack()
        e_per = ctk.CTkEntry(v, width=200); e_per.insert(0, m_per); e_per.pack(pady=5)
        
        ctk.CTkLabel(v, text="Puntos Ranking:").pack()
        e_pts = ctk.CTkEntry(v, width=200); e_pts.insert(0, p_pts); e_pts.pack(pady=5)
        
        def save():
            try:
                old_price = p_cli
                new_price = e_cli.get()
                new_margen = e_per.get()
                new_pts = e_pts.get()
                
                if actualizar_precio_db(div, float(new_price), float(new_margen), int(new_pts)):
                    registrar_log("CAMBIO_TARIFA", f"División {div}: ${old_price} -> ${new_price} | Pts: {new_pts}")
                    v.destroy(); self.actualizar_tabla_precios()
            except ValueError:
                messagebox.showerror("Error", "Valores numéricos inválidos.")

        ctk.CTkButton(v, text="Actualizar", command=save, fg_color="#3498db").pack(pady=20)
    
    def eliminar_precio_seleccionado(self):
        sel = self.tabla_precios.selection()
        if sel and messagebox.askyesno("Confirmar", "¿Eliminar tarifa?", parent=self):
            if eliminar_precio_db(self.tabla_precios.item(sel)['values'][0]): self.actualizar_tabla_precios()

    def ejecutar_backup_manual(self):
        try:
            realizar_backup_db()
            messagebox.showinfo("Seguridad", "✅ Base de datos respaldada con éxito.\nRevisa la carpeta /backups")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el respaldo: {e}")

    def abrir_visor_logs(self):

        v = ctk.CTkToplevel(self)
        v.title("Registro de Seguridad y Auditoría")
        self.centrar_ventana(v, 700, 500)
        v.attributes("-topmost", True)
        v.grab_set()

        header = ctk.CTkFrame(v, fg_color="transparent")
        header.pack(fill="x", padx=20, pady=15)
        ctk.CTkLabel(header, text="🛡️ LOGS DE SISTEMA (CAJA NEGRA)", font=("Consolas", 16, "bold")).pack(side="left")
        ctk.CTkButton(header, text="Cerrar", width=80, fg_color="#e74c3c", command=v.destroy).pack(side="right")

        txt_frame = ctk.CTkFrame(v, fg_color="#1a1a1a", corner_radius=10)
        txt_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        txt = ctk.CTkTextbox(txt_frame, width=600, height=400, font=("Consolas", 11), fg_color="#0f0f0f", text_color="#00ff00", activate_scrollbars=True)
        txt.pack(fill="both", expand=True, padx=5, pady=5)

        logs = obtener_logs_db(limite=100)

        if not logs:
            txt.insert("0.0", "\n   [ SYSTEM ] No hay registros de actividad reciente.")
        else:
            header_txt = f"{'FECHA/HORA':<20} | {'EVENTO':<18} | DETALLES\n"
            sep = "-"*85 + "\n"
            txt.insert("0.0", header_txt + sep)

            for log in logs:
                fecha_corta = log[0][5:-3]
                linea = f"{fecha_corta:<20} | {log[1]:<18} | {log[2]}\n"
                txt.insert("end", linea)
        txt.configure(state="disabled")

    def on_closing(self):
        print("💾 Realizando backup automático antes de cerrar...")
        try:
            realizar_backup_db()
        except Exception as e:
            print(f"⚠️ Error en backup al cerrar: {e}")
        try:
            plt.close('all')
        except:
            pass
        self.destroy()
        print("🔴 Apagado forzoso del sistema.")
        os._exit(0)

# =============================================================================
# EJECUCIÓN PRINCIPAL
# =============================================================================

if __name__ == "__main__":
    app = PerezBoostApp()
    app.mainloop()